"""
보존력 시험 OCR 백엔드 로직
Streamlit에서 직접 import하여 사용
"""

import io
import re
import fitz
import requests
import pandas as pd
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
from openpyxl import Workbook
import os
import logging
import math
from typing import List, Dict, Tuple, Optional

# 로깅 설정
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

from dotenv import load_dotenv
load_dotenv()
# 설정
UPSTAGE_API_KEY = os.getenv("UPSTAGE_API_KEY")
UPSTAGE_URL = "https://api.upstage.ai/v1/document-ai/document-parse"
STRAINS = ['E.coli', 'P.aeruginosa', 'S.aureus', 'C.albicans', 'A.brasiliensis']


class PDFProcessor:
    """PDF 처리 클래스"""
    
    @staticmethod
    def extract_page_count(pdf_bytes: bytes) -> int:
        """PDF 페이지 수 추출"""
        try:
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")
            return doc.page_count
        except Exception as e:
            logger.error(f"페이지 수 추출 실패: {e}")
            return 0
    
    @staticmethod
    def render_page_image(pdf_bytes: bytes, page_index: int, zoom: float = 2.0) -> bytes:
        """PDF 페이지를 이미지로 렌더링"""
        try:
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")
            page = doc.load_page(page_index)
            mat = fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat, alpha=False)
            return pix.tobytes("png")
        except Exception as e:
            logger.error(f"이미지 렌더링 실패: {e}")
            return None


class OCRProcessor:
    """OCR 처리 클래스"""
    
    @staticmethod
    def request_ocr(image_bytes: bytes) -> Optional[dict]:
        """업스테이지 OCR API 호출"""
        try:
            headers = {"Authorization": f"Bearer {UPSTAGE_API_KEY}"}
            files = {"document": ("image.jpg", image_bytes, "image/jpeg")}
            data = {
                "model": "document-parse",
                "ocr": "force",
                "base64_encoding": "['table']"
            }
            
            response = requests.post(
                UPSTAGE_URL, 
                headers=headers, 
                files=files, 
                data=data, 
                timeout=120
            )
            
            if response.status_code == 200:
                return response.json()
            else:
                logger.error(f"OCR API 오류: {response.status_code}")
                return None
                
        except Exception as e:
            logger.error(f"OCR 요청 실패: {e}")
            return None
    
    @staticmethod
    def parse_table_from_ocr(ocr_result: dict) -> Tuple[List[dict], dict]:
        """OCR 결과에서 테이블 파싱"""
        try:
            html_parts = []
            if 'elements' in ocr_result:
                for element in ocr_result.get("elements", []):
                    content = element.get("content", {})
                    html = content.get("html", "")
                    if html:
                        html_parts.append(html)
            
            if not html_parts:
                return [], {}
            
            html_content = "<html><body>\n" + "\n".join(html_parts) + "\n</body></html>"
            soup = BeautifulSoup(html_content, 'html.parser')
            table = soup.find('table')
            
            if not table:
                return [], {}
            
            rows = table.find_all('tr')
            if len(rows) < 3:
                return [], {}
            
            # 날짜 정보 추출
            date_info = DataCleaner.extract_date_info(rows)
            
            # 테이블 데이터 파싱
            table_data = DataCleaner.parse_table_rows(rows)
            
            return table_data, date_info
            
        except Exception as e:
            logger.error(f"테이블 파싱 오류: {e}")
            return [], {}


class DataCleaner:
    """데이터 정제 클래스"""
    
    @staticmethod
    def extract_date_info(rows) -> dict:
        """
        날짜 정보 추출 (개선 버전)
        
        개선 사항:
        - 연속된 날짜 문자열 지원 추가
        - 기존 로직 유지
        """
        date_info = {}
        if len(rows) >= 2:
            header_cells = rows[1].find_all('td')
            if len(header_cells) >= 1:
                first_date_str = header_cells[0].text.strip()
                
                # 🆕 연속 날짜 패턴 먼저 시도
                consecutive_dates = DataCleaner.parse_consecutive_dates(first_date_str)
                if consecutive_dates and len(consecutive_dates) >= 4:
                    date_info = {
                        'date_0': consecutive_dates[0],
                        'date_7': consecutive_dates[1],
                        'date_14': consecutive_dates[2],
                        'date_28': consecutive_dates[3]
                    }
                    return date_info
                
                # 기존 방식 (단일 날짜 파싱)
                first_date = DataCleaner.parse_date(first_date_str)
                
                if first_date:
                    date_info = {
                        'date_0': first_date.strftime("%m/%d"),
                        'date_7': (first_date + timedelta(days=7)).strftime("%m/%d"),
                        'date_14': (first_date + timedelta(days=14)).strftime("%m/%d"),
                        'date_28': (first_date + timedelta(days=28)).strftime("%m/%d")
                    }
        return date_info
    
    @staticmethod
    def parse_table_rows(rows) -> List[dict]:
        """테이블 행 파싱"""
        table_data = []
        current_test_number = None
        current_prescription_number = None
        
        for i, row in enumerate(rows[2:], start=3):
            cells = row.find_all('td')
            if len(cells) < 2:
                continue
            
            # Bulk Name 행 감지
            if cells[0].get('rowspan') and cells[0].text.strip():
                bulk_name = cells[0].text.strip()
                current_test_number, current_prescription_number = DataCleaner.extract_numbers(bulk_name)
                
                if len(cells) > 1:
                    strain = cells[1].text.strip()
                    cfu_indices = {'0일': 3, '7일': 4, '14일': 5, '28일': 6, '판정': 7, '최종판정': 8}
                else:
                    continue
            else:
                if len(cells) < 1:
                    continue
                strain = cells[0].text.strip()
                cfu_indices = {'0일': 2, '7일': 3, '14일': 4, '28일': 5, '판정': 6, '최종판정': 7}
            
            # 유효한 균주 확인
            valid_strains = STRAINS + ['Escherichia', 'Pseudomonas', 'Staphylococcus', 'Candida', 'Aspergillus']
            if not strain or not any(valid_strain in strain for valid_strain in valid_strains):
                continue
            
            strain_normalized = DataCleaner.normalize_strain_name(strain)
            
            # CFU 데이터 추출
            row_data = {
                'test_number': current_test_number or '',
                'prescription_number': current_prescription_number or '',
                'strain': strain_normalized,
                'cfu_0day': DataCleaner.clean_cfu_value(
                    cells[cfu_indices['0일']].text.strip() if len(cells) > cfu_indices['0일'] else "", 
                    strain_normalized, '0일'
                ),
                'cfu_7day': DataCleaner.clean_cfu_value(
                    cells[cfu_indices['7일']].text.strip() if len(cells) > cfu_indices['7일'] else "", 
                    strain_normalized, '7일'
                ),
                'cfu_14day': DataCleaner.clean_cfu_value(
                    cells[cfu_indices['14일']].text.strip() if len(cells) > cfu_indices['14일'] else "", 
                    strain_normalized, '14일'
                ),
                'cfu_28day': DataCleaner.clean_cfu_value(
                    cells[cfu_indices['28일']].text.strip() if len(cells) > cfu_indices['28일'] else "", 
                    strain_normalized, '28일'
                ),
                'judgment': DataCleaner.get_judgment_value(cells, cfu_indices),
                'final_judgment': DataCleaner.get_final_judgment_value(cells, cfu_indices)
            }
            
            if any(v for k, v in row_data.items() if k.startswith('cfu_') and v.strip()):
                table_data.append(row_data)
        
        return table_data
    
    @staticmethod
    def extract_numbers(bulk_name: str) -> Tuple[Optional[str], Optional[str]]:
        """
        시험번호와 처방번호 추출 (개선 버전)
        
        개선 사항:
        - A-L 범위로 확장 (기존: A-Z)
        - I/1 OCR 오류 자동 보정
        - 더 많은 처방번호 패턴 지원
        - 공백 처리 강화
        """
        test_number = None
        prescription_number = None
        
        try:
            # 전처리
            bulk_name = bulk_name.upper()
            bulk_name = bulk_name.replace('!', 'I')  # OCR 오류 보정
            bulk_name = re.sub(r'-\s+', '-', bulk_name)  # '- ' → '-'
            bulk_name = re.sub(r'\s+', ' ', bulk_name)   # 연속 공백 제거
            
            # ======== 처방번호 패턴 (확장) ========
            prescription_patterns = [
                r'\b[A-Z]{2,4}\d{4,5}[A-Z]?-[A-Z]{1,5}\d?\b',
                r'\b[A-Z]{3}\d{5}-[A-Z]{2,4}\b',
                r'\b[A-Z]{2,4}\d{3,6}-[A-Z]{1,5}\b',
                r'\b[A-Z]{2,5}\d{4}-[A-Z]{1,3}\d{0,2}\b',
                r'\b[A-Z]{2,4}\d{4,5}[A-Z]?-\s*[A-Z]{1,5}\d?\b',  # 공백 허용
                r'\b[A-Z]{2,4}\d{3,5}-[A-Z]{1,4}\d{1,2}\b',  # AZLY1 타입
            ]
            
            all_prescription_matches = []
            for pattern in prescription_patterns:
                matches = re.findall(pattern, bulk_name)
                all_prescription_matches.extend(matches)
            
            # ======== 시험번호 패턴 (A-L 확장 + OCR 보정) ========
            all_test_matches = []
            
            # 정상 형태 (I가 정확히 인식된 경우)
            correct_matches = re.findall(r'\b(\d{2}[A-L]\d{2}I\d{2,3})\b', bulk_name)
            all_test_matches.extend(correct_matches)
            
            # OCR 오류 형태 (I를 1로 잘못 인식)
            ocr_error_patterns = [
                r'\b(\d{2}[A-L]\d{2}1\d{2,3})\b',   # I가 1로
                r'\b(\d{2}[A-L]\d{5,6})\b',         # I 누락
            ]
            
            for pattern in ocr_error_patterns:
                matches = re.findall(pattern, bulk_name)
                for match in matches:
                    if len(match) == 7:  # 25A2012 → 25A20I2
                        corrected = match[:5] + 'I' + match[6:]
                        all_test_matches.append(corrected)
                        logger.info(f"OCR I/1 보정: '{match}' → '{corrected}'")
                    elif len(match) == 8:  # 25A20102 → 25A20I02
                        corrected = match[:5] + 'I' + match[6:]
                        all_test_matches.append(corrected)
                        logger.info(f"OCR I 삽입 보정: '{match}' → '{corrected}'")
            
            # 공백이 있는 형태 (A-L 확장)
            raw_matches = re.findall(r'(\d{2})([A-L])(\d)\s+(\d)(\d{2,3})', bulk_name)
            for year_prefix, letter, d1, d2, last_digits in raw_matches:
                converted = f"{year_prefix}{letter}{d1}{d2}I{last_digits[:2]}"
                all_test_matches.append(converted)
            
            # 중복 제거
            all_test_matches = list(dict.fromkeys(all_test_matches))
            all_prescription_matches = list(dict.fromkeys(all_prescription_matches))
            
            test_number = all_test_matches[0] if all_test_matches else None
            prescription_number = all_prescription_matches[0] if all_prescription_matches else None
            
            return test_number, prescription_number
            
        except Exception as e:
            logger.warning(f"번호 추출 중 오류: {e}")
            return None, None
    
    @staticmethod
    def normalize_strain_name(strain: str) -> str:
        """균주명 정규화"""
        strain_mapping = {
            'E.coli': 'E.coli', 'Escherichia coli': 'E.coli', 'E. coli': 'E.coli',
            'P.aeruginosa': 'P.aeruginosa', 'Pseudomonas aeruginosa': 'P.aeruginosa', 'P. aeruginosa': 'P.aeruginosa',
            'S.aureus': 'S.aureus', 'Staphylococcus aureus': 'S.aureus', 'S. aureus': 'S.aureus',
            'C.albicans': 'C.albicans', 'Candida albicans': 'C.albicans', 'C. albicans': 'C.albicans',
            'A.brasiliensis': 'A.brasiliensis', 'Aspergillus brasiliensis': 'A.brasiliensis', 'A. brasiliensis': 'A.brasiliensis'
        }
        
        for full_name, short_name in strain_mapping.items():
            if full_name.lower() == strain.lower():
                return short_name
        
        for full_name, short_name in strain_mapping.items():
            if full_name.lower() in strain.lower():
                return short_name
        
        return strain
    
    @staticmethod
    def clean_cfu_value(value: str, strain: str = None, day_column: str = None) -> str:
        """CFU 값 정리 및 보정"""
        if not value:
            return ""
        
        original_value = value
        
        # OCR 오류 제거
        value = re.sub(r'[ぁ-んァ-ン一-龯]+', '', value)
        value = value.replace('く', '<').replace('C', '<').replace('O', '0')
        value = value.replace('Co', '0').replace('CIO', '<10').replace('C10', '<10')
        value = value.strip()
        
        # 지수 형태 처리
        if re.search(r'[×xX]', value):
            exp_match = re.match(r'([0-9.]+)\s*[×xX]\s*10\s*\^?([0-9]+)', value)
            if exp_match:
                base = exp_match.group(1)
                exp = exp_match.group(2)
                return f"{base}×10^{exp}"
        
        # <10 형태 처리
        if '<' in value:
            if re.search(r'<\s*10\s*\^?\s*([0-9]+)', value):
                exp = re.search(r'<\s*10\s*\^?\s*([0-9]+)', value).group(1)
                return f"<10^{exp}"
            elif re.search(r'<\s*([0-9]+)', value):
                return f"<{re.search(r'<\s*([0-9]+)', value).group(1)}"
            return "<10"
        
        # ≤ 형태 처리
        if '≤' in value:
            if re.search(r'≤\s*([0-9]+)', value):
                num = re.search(r'≤\s*([0-9]+)', value).group(1)
                return f"≤{num}"
        
        # 균주별 보정
        target_strains = ['E.coli', 'P.aeruginosa', 'S.aureus', 'C.albicans']
        is_target_strain = strain and any(s in strain for s in target_strains)
        
        if day_column in ['7일', '14일', '28일'] and is_target_strain:
            preserve_patterns = [r'^≤\d+[°⁰]?$']
            should_preserve = any(re.match(pattern, value, re.IGNORECASE) for pattern in preserve_patterns)
            if should_preserve:
                return value
            
            if len(original_value) >= 6:
                return value
            
            if day_column == '7일':
                corrected_value = "<10^2"
            elif day_column in ['14일', '28일']:
                corrected_value = "<10"
            else:
                corrected_value = "<10"
            
            has_clear_power_signal = ('2' in original_value and 
                                    any(char in original_value for char in ['^', '²', '⁰', '¹', '²', '³']))
            
            if has_clear_power_signal and day_column != '28일':
                corrected_value = "<10^2"
            
            return corrected_value
        
        return value
    
    @staticmethod
    def get_judgment_value(cells, cfu_indices: dict) -> str:
        """판정 값 추출"""
        try:
            if len(cells) > cfu_indices['판정']:
                raw_value = cells[cfu_indices['판정']].text.strip()
                if any(char in raw_value for char in ['X', '×', 'v', 'V']):
                    return '부적합'
                return '적합'
            return "적합"
        except:
            return "적합"
    
    @staticmethod
    def get_final_judgment_value(cells, cfu_indices: dict) -> str:
        """최종판정 값 추출"""
        try:
            if len(cells) > cfu_indices['최종판정']:
                raw_value = cells[cfu_indices['최종판정']].text.strip()
                if any(char in raw_value for char in ['X', '×', 'v', 'V']):
                    return '부적합'
                return '적합'
            return "적합"
        except:
            return "적합"
        
    @staticmethod
    def parse_consecutive_dates(date_text: str) -> List[str]:
        """
        연속된 날짜 문자열 파싱
        
        예시: '01 15 01 22 01 29 02 12' → ['01/15', '01/22', '01/29', '02/12']
        
        Args:
            date_text (str): 연속된 날짜 문자열
            
        Returns:
            List[str]: 날짜 리스트 (4개)
        """
        try:
            parts = date_text.split()
            
            if len(parts) >= 8 and all(part.isdigit() and len(part) == 2 for part in parts):
                dates = []
                for i in range(0, min(8, len(parts)), 2):
                    if i + 1 < len(parts):
                        month = parts[i]
                        day = parts[i + 1]
                        dates.append(f"{month}/{day}")
                
                if len(dates) >= 4:
                    return dates[:4]
            
            return []
            
        except Exception as e:
            logger.warning(f"연속 날짜 파싱 오류: {e}")
            return []
        
    @staticmethod
    def parse_date(date_str: str) -> Optional[datetime]:
        """날짜 문자열을 datetime 객체로 변환"""
        try:
            date_formats = [
                '%m %d', '%m-%d', '%m/%d', '%m.%d',
                '%m월%d일', '%m월 %d일',
                '%d/%m', '%d-%m', '%d %m'
            ]
            
            for date_format in date_formats:
                try:
                    return datetime.strptime(date_str, date_format)
                except ValueError:
                    continue
            
            if re.match(r'^\d+\s+\d+$', date_str):
                try:
                    return datetime.strptime(date_str, '%m %d')
                except ValueError:
                    pass
            
            return None
        except:
            return None
    
    @staticmethod
    def convert_to_log(cfu_value: str) -> str:
        """CFU → Log 변환"""
        if not cfu_value:
            return ""
        
        try:
            if '<' in cfu_value:
                if '10^' in cfu_value:
                    exp_match = re.search(r'<10\^(\d+)', cfu_value)
                    if exp_match:
                        return f"<{exp_match.group(1)}.0"
                elif '≤' in cfu_value:
                    num_match = re.search(r'≤(\d+)', cfu_value)
                    if num_match:
                        return f"<{num_match.group(1)}.0"
                return "<1.0"
            
            exp_match = re.match(r'([0-9.]+)×10\^(\d+)', cfu_value)
            if exp_match:
                base = float(exp_match.group(1))
                exp = int(exp_match.group(2))
                log_value = exp + math.log10(base)
                return round(log_value, 1)
            
            try:
                num = float(cfu_value)
                return round(math.log10(num), 1)
            except ValueError:
                pass
            
            return cfu_value
            
        except Exception as e:
            logger.warning(f"Log 변환 실패: {cfu_value}, 오류: {e}")
            return cfu_value


class ExcelIncrementalSaver:
    """
    Excel 증분 저장 관리 클래스
    
    기능:
    - 페이지 처리할 때마다 즉시 Excel 파일에 저장
    - 템플릿 기반 시트 생성 (copy_worksheet 사용)
    - 중복 시트명 자동 처리
    """
    
    # 🆕 기본 템플릿 파일 경로
    DEFAULT_TEMPLATE = "TestResult_OCR_v1.xlsx"
    
    def __init__(self, output_path="보존력시험_최종.xlsx", template_file=None):
        """
        Args:
            output_path (str): 저장할 Excel 파일 경로
            template_file (str): 템플릿 Excel 파일 경로 (None이면 기본값 사용)
        """
        self.output_path = output_path
        
        # 🆕 template_file이 None이면 기본 템플릿 사용
        if template_file is None:
            self.template_file = self.DEFAULT_TEMPLATE
        else:
            self.template_file = template_file
        
        # 🆕 템플릿 파일 존재 확인
        if not os.path.exists(self.template_file):
            logger.warning(f"⚠️ 템플릿 파일을 찾을 수 없습니다: {self.template_file}")
            logger.warning("빈 Excel 파일로 생성됩니다.")
            self.template_file = None
        else:
            logger.info(f"✅ 템플릿 파일 확인: {self.template_file}")
        
        # 파일이 없으면 초기화
        if not os.path.exists(self.output_path):
            self._initialize_excel()
    
    def _initialize_excel(self):
        """Excel 파일 초기화"""
        try:
            if self.template_file and os.path.exists(self.template_file):
                # 🆕 템플릿 파일 전체 복사
                import shutil
                shutil.copy2(self.template_file, self.output_path)
                
                # 🆕 첫 번째 시트를 TEMPLATE_BASE로 이름 변경
                from openpyxl import load_workbook
                workbook = load_workbook(self.output_path)
                
                if len(workbook.sheetnames) > 0:
                    first_sheet = workbook[workbook.sheetnames[0]]
                    first_sheet.title = "TEMPLATE_BASE"
                    logger.info(f"✅ 템플릿 시트 '{workbook.sheetnames[0]}' → 'TEMPLATE_BASE'로 변경")
                
                workbook.save(self.output_path)
                workbook.close()
                
                logger.info(f"✅ 템플릿 기반 Excel 초기화 완료: {self.output_path}")
            else:
                # 빈 Excel 생성
                wb = Workbook()
                wb.remove(wb.active)
                wb.save(self.output_path)
                wb.close()
                
                logger.warning(f"⚠️ 템플릿 없이 빈 Excel 파일 생성: {self.output_path}")
            
            return True
        except Exception as e:
            logger.error(f"❌ Excel 초기화 실패: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def add_test_data(self, test_data, date_info=None):
        """
        테스트 데이터를 Excel에 추가
        
        Args:
            test_data: DataFrame 또는 딕셔너리 리스트
            date_info: 날짜 정보 딕셔너리 또는 리스트
            
        Returns:
            bool: 성공 여부
        """
        try:
            from openpyxl import load_workbook
            
            # Excel 파일 로드
            workbook = load_workbook(self.output_path)
            
            # DataFrame으로 변환
            if isinstance(test_data, pd.DataFrame):
                df = test_data
            elif isinstance(test_data, list):
                df = pd.DataFrame(test_data)
            else:
                logger.error("❌ 지원하지 않는 데이터 형식")
                return False
            
            if df.empty:
                logger.warning("⚠️ 빈 데이터 - 저장 건너뛰기")
                return False
            
            # 시험번호 추출
            test_number = df.iloc[0].get('test_number', '')
            if not test_number:
                test_number = "Unknown_Test"
                logger.warning(f"⚠️ 시험번호 없음 - 기본값 사용: {test_number}")
            
            # 중복 시트명 처리
            sheet_name = str(test_number)
            counter = 1
            original_name = sheet_name
            while sheet_name in workbook.sheetnames:
                sheet_name = f"{original_name}_{counter}"
                counter += 1
            
            # 🆕 템플릿 시트 복사하여 새 시트 생성
            if "TEMPLATE_BASE" in workbook.sheetnames:
                template_sheet = workbook["TEMPLATE_BASE"]
                new_sheet = workbook.copy_worksheet(template_sheet)
                new_sheet.title = sheet_name
                logger.info(f"✅ 템플릿 시트 복사 완료: {sheet_name}")
            else:
                # 템플릿이 없으면 빈 시트 생성
                new_sheet = workbook.create_sheet(title=sheet_name)
                logger.warning(f"⚠️ 템플릿 없이 빈 시트 생성: {sheet_name}")
            
            # 데이터 매핑
            self._map_data_to_sheet(new_sheet, df, date_info)
            
            # 즉시 저장
            workbook.save(self.output_path)
            workbook.close()
            
            logger.info(f"💾 Excel 저장 완료: {sheet_name} 시트 추가")
            return True
            
        except Exception as e:
            logger.error(f"❌ Excel 저장 실패: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _map_data_to_sheet(self, worksheet, df, date_info):
        """데이터를 시트에 매핑"""
        try:
            if df.empty:
                logger.warning("⚠️ 빈 DataFrame - 매핑 건너뛰기")
                return
            
            # 시험번호 매핑
            test_number = df.iloc[0].get('test_number', '')
            worksheet['AA3'] = test_number  # 원본 보고서
            worksheet['AA33'] = test_number  # Log 보고서
            logger.info(f"📝 시험번호 매핑: AA3, AA33 = {test_number}")
            
            # 처방번호 매핑
            if 'prescription_number' in df.columns:
                prescription_number = df.iloc[0].get('prescription_number', '')
                if prescription_number:
                    worksheet['E4'] = prescription_number  # 원본
                    worksheet['E34'] = prescription_number  # Log
                    logger.info(f"📝 처방번호 매핑: E4, E34 = {prescription_number}")
            
            # 날짜 정보 매핑
            if date_info:
                # 딕셔너리인 경우
                if isinstance(date_info, dict):
                    date_list = [
                        date_info.get('date_0', ''),
                        date_info.get('date_7', ''),
                        date_info.get('date_14', ''),
                        date_info.get('date_28', '')
                    ]
                # 리스트인 경우
                elif isinstance(date_info, list):
                    date_list = date_info
                else:
                    date_list = []
                
                if len(date_list) >= 4:
                    date_positions_original = ['I19', 'L19', 'O19', 'R19']
                    date_positions_log = ['I49', 'L49', 'O49', 'R49']
                    
                    for i, date_val in enumerate(date_list[:4]):
                        if date_val:  # 빈 값이 아닌 경우만 매핑
                            worksheet[date_positions_original[i]] = date_val
                            worksheet[date_positions_log[i]] = date_val
                    
                    logger.info(f"📅 날짜 정보 매핑: {date_list}")
            
            # 균주별 CFU 데이터 매핑
            strain_mapping = {
                'E.coli': 'E.coli',
                'Escherichia coli': 'E.coli',
                'P.aeruginosa': 'P.aeruginosa',
                'Pseudomonas aeruginosa': 'P.aeruginosa',
                'S.aureus': 'S.aureus',
                'Staphylococcus aureus': 'S.aureus',
                'C.albicans': 'C.albicans',
                'Candida albicans': 'C.albicans',
                'A.brasiliensis': 'A.brasiliensis',
                'Aspergillus brasiliensis': 'A.brasiliensis'
            }
            
            original_positions = {
                'E.coli': ['J20', 'M20', 'P20', 'S20', 'U20'],
                'P.aeruginosa': ['J21', 'M21', 'P21', 'S21', 'U21'],
                'S.aureus': ['J22', 'M22', 'P22', 'S22', 'U22'],
                'C.albicans': ['J23', 'M23', 'P23', 'S23', 'U23'],
                'A.brasiliensis': ['J24', 'M24', 'P24', 'S24', 'U24']
            }
            
            log_positions = {
                'E.coli': ['J50', 'M50', 'P50', 'S50'],
                'P.aeruginosa': ['J51', 'M51', 'P51', 'S51'],
                'S.aureus': ['J52', 'M52', 'P52', 'S52'],
                'C.albicans': ['J53', 'M53', 'P53', 'S53'],
                'A.brasiliensis': ['J54', 'M54', 'P54', 'S54']
            }
            
            mapped_count = 0
            for _, row in df.iterrows():
                strain = row.get('strain', '')
                if not strain:
                    continue
                
                mapped_strain = strain_mapping.get(strain, strain)
                
                if mapped_strain in original_positions:
                    # 원본 CFU 값
                    positions = original_positions[mapped_strain]
                    worksheet[positions[0]] = row.get('cfu_0day', '')
                    worksheet[positions[1]] = row.get('cfu_7day', '')
                    worksheet[positions[2]] = row.get('cfu_14day', '')
                    worksheet[positions[3]] = row.get('cfu_28day', '')
                    worksheet[positions[4]] = row.get('judgment', '')
                    
                    # Log 값
                    log_pos = log_positions[mapped_strain]
                    worksheet[log_pos[0]] = DataCleaner.convert_to_log(row.get('cfu_0day', ''))
                    worksheet[log_pos[1]] = DataCleaner.convert_to_log(row.get('cfu_7day', ''))
                    worksheet[log_pos[2]] = DataCleaner.convert_to_log(row.get('cfu_14day', ''))
                    worksheet[log_pos[3]] = DataCleaner.convert_to_log(row.get('cfu_28day', ''))
                    
                    mapped_count += 1
                    logger.info(f"🦠 {mapped_strain} 데이터 매핑 완료")
            
            logger.info(f"✅ 총 {mapped_count}개 균주 데이터 매핑 완료")
            
        except Exception as e:
            logger.error(f"❌ 데이터 매핑 실패: {e}")
            import traceback
            traceback.print_exc()
    
    def get_sheet_list(self):
        """현재 Excel 파일의 시트 목록 반환"""
        try:
            from openpyxl import load_workbook
            
            if os.path.exists(self.output_path):
                workbook = load_workbook(self.output_path, read_only=True)
                sheet_names = workbook.sheetnames
                workbook.close()
                
                # TEMPLATE_BASE 제외
                filtered_names = [name for name in sheet_names if name != "TEMPLATE_BASE"]
                logger.info(f"📋 시트 목록: {filtered_names}")
                return filtered_names
            else:
                logger.warning(f"⚠️ Excel 파일이 존재하지 않음: {self.output_path}")
                return []
        except Exception as e:
            logger.error(f"❌ 시트 목록 조회 실패: {e}")
            return []
    
    def get_excel_bytes(self):
        """Excel 파일을 바이트로 읽어서 반환 (다운로드용)"""
        try:
            if os.path.exists(self.output_path):
                with open(self.output_path, 'rb') as f:
                    excel_bytes = f.read()
                logger.info(f"✅ Excel 파일 읽기 완료: {len(excel_bytes)} bytes")
                return excel_bytes
            else:
                logger.warning(f"⚠️ Excel 파일이 존재하지 않음: {self.output_path}")
                return None
        except Exception as e:
            logger.error(f"❌ Excel 읽기 실패: {e}")
            return None
    
    def get_statistics(self):
        """Excel 파일 통계 정보 반환"""
        try:
            from openpyxl import load_workbook
            
            if not os.path.exists(self.output_path):
                return {
                    'total_sheets': 0,
                    'test_sheets': 0,
                    'file_size': 0
                }
            
            workbook = load_workbook(self.output_path, read_only=True)
            total_sheets = len(workbook.sheetnames)
            test_sheets = len([name for name in workbook.sheetnames if name != "TEMPLATE_BASE"])
            workbook.close()
            
            file_size = os.path.getsize(self.output_path)
            
            stats = {
                'total_sheets': total_sheets,
                'test_sheets': test_sheets,
                'file_size': file_size,
                'file_size_mb': round(file_size / (1024 * 1024), 2)
            }
            
            logger.info(f"📊 통계: {stats}")
            return stats
            
        except Exception as e:
            logger.error(f"❌ 통계 조회 실패: {e}")
            return {
                'total_sheets': 0,
                'test_sheets': 0,
                'file_size': 0
            }

# 편의 함수
def process_pdf_page(pdf_bytes: bytes, page_index: int) -> dict:
    """PDF 페이지 전체 처리 파이프라인"""
    result = {
        'success': False,
        'data': [],
        'date_info': {},
        'message': ''
    }
    
    try:
        # 1. 이미지 렌더링
        img_bytes = PDFProcessor.render_page_image(pdf_bytes, page_index)
        if not img_bytes:
            result['message'] = "이미지 렌더링 실패"
            return result
        
        # 2. OCR 처리
        ocr_result = OCRProcessor.request_ocr(img_bytes)
        if not ocr_result:
            result['message'] = "OCR 처리 실패"
            return result
        
        # 3. 테이블 파싱
        table_data, date_info = OCRProcessor.parse_table_from_ocr(ocr_result)
        
        result['success'] = True
        result['data'] = table_data
        result['date_info'] = date_info
        result['message'] = f"{len(table_data)}개 균주 데이터 추출 완료"
        
        return result
        
    except Exception as e:
        logger.error(f"처리 오류: {e}")
        result['message'] = str(e)
        return result