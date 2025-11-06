"""
보존력 시험 OCR 백엔드 (Azure Document Intelligence 기반)
스킨케어 팀 구조를 참고하여 보존력 시험에 맞게 커스터마이징
"""

import io
import logging
import os
import tempfile
import re
from typing import List, Dict, Tuple, Optional
from datetime import datetime, timedelta
import math

# Azure OCR
from azure.ai.formrecognizer import DocumentAnalysisClient
from azure.core.credentials import AzureKeyCredential

# 기존 backend에서 PDFProcessor만 import
from backend import PDFProcessor

# Excel 처리
try:
    import openpyxl
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️ openpyxl 라이브러리를 찾을 수 없습니다. pip install openpyxl로 설치하세요.")

# ========================================
# 로그 설정 (파일 + 콘솔 동시 출력)
# ========================================
def setup_logging():
    """로그 파일 및 콘솔 출력 설정"""
    
    # 로그 디렉토리 생성
    log_dir = "logs"
    if not os.path.exists(log_dir):
        os.makedirs(log_dir)
    
    # 로그 파일명 (타임스탬프 포함)
    log_filename = os.path.join(
        log_dir, 
        f"preservation_ocr_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    )
    
    # 로거 설정
    logger = logging.getLogger(__name__)
    logger.setLevel(logging.DEBUG)  # 모든 레벨 기록
    
    # 🔧 중복 출력 방지: 상위 로거로 전파 차단
    logger.propagate = False
    
    # 기존 핸들러 제거 (중복 방지)
    if logger.hasHandlers():
        logger.handlers.clear()
    
    # 포맷 설정
    formatter = logging.Formatter(
        '%(asctime)s | %(levelname)-8s | %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    
    # 1. 파일 핸들러 (모든 로그를 파일에 저장)
    file_handler = logging.FileHandler(log_filename, encoding='utf-8')
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)
    
    # 2. 콘솔 핸들러 (INFO 이상만 콘솔 출력)
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(formatter)
    logger.addHandler(console_handler)
    
    # 초기 메시지
    logger.info("="*80)
    logger.info("🚀 보존력 시험 OCR 시스템 시작")
    logger.info(f"📁 로그 파일: {log_filename}")
    logger.info("="*80)
    
    return logger

# 로거 초기화
logger = setup_logging()

# 환경 변수
AZURE_KEY = os.getenv('AZURE_KEY', '')
AZURE_ENDPOINT = os.getenv('AZURE_ENDPOINT', '')

# 균주 목록
STRAINS = ['E.coli', 'P.aeruginosa', 'S.aureus', 'C.albicans', 'A.brasiliensis']


def load_progress_excel(excel_path: str) -> Dict[str, Dict[str, str]]:
    """
    TestResult_PROGRESS.xlsx 파일에서 추가 정보 로드
    
    Args:
        excel_path: Excel 파일 경로
        
    Returns:
        처방번호를 키로 하는 딕셔너리
        {
            'WC1982-D1FK': {
                'product_name': 'SRM-VPHCM(F)DU 조제',
                'formulation': 'O/W',
                'preservative_info': '1,2-HEXANEDIOL 0.9%HYDROXYACETOP ISO'
            },
            ...
        }
    """
    if not OPENPYXL_AVAILABLE:
        logger.warning("⚠️ openpyxl이 설치되지 않아 진행서를 읽을 수 없습니다.")
        return {}
    
    if not os.path.exists(excel_path):
        logger.warning(f"⚠️ 진행서 파일을 찾을 수 없습니다: {excel_path}")
        return {}
    
    try:
        logger.info(f"\n📖 진행서 파일 읽기: {excel_path}")
        
        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active
        
        # 데이터 딕셔너리 초기화
        data_dict = {}
        
        # 헤더 행 찾기 (첫 번째 행 또는 "제품명" 포함 행)
        header_row = 1
        for row_idx in range(1, min(10, ws.max_row + 1)):
            cell_value = ws.cell(row_idx, 1).value
            if cell_value and '제품명' in str(cell_value):
                header_row = row_idx
                break
        
        logger.info(f"  📍 헤더 행: {header_row}")
        
        # 컬럼 인덱스 찾기
        col_map = {}
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(header_row, col_idx).value
            if cell_value:
                cell_value_str = str(cell_value).strip()
                if '제품명' in cell_value_str:
                    col_map['product_name'] = col_idx
                elif '처방번호' in cell_value_str:
                    col_map['prescription'] = col_idx
                elif '제형' in cell_value_str:
                    col_map['formulation'] = col_idx
                elif '고시미등록방부보조' in cell_value_str:
                    col_map['preservative'] = col_idx
        
        logger.info(f"  📋 컬럼 매핑: {col_map}")
        
        # 필수 컬럼 확인
        required_cols = ['prescription', 'product_name', 'formulation', 'preservative']
        missing_cols = [col for col in required_cols if col not in col_map]
        if missing_cols:
            logger.warning(f"  ⚠️ 필수 컬럼 누락: {missing_cols}")
            return {}
        
        # 데이터 행 읽기
        data_count = 0
        for row_idx in range(header_row + 1, ws.max_row + 1):
            # 처방번호 읽기
            prescription = ws.cell(row_idx, col_map['prescription']).value
            if not prescription:
                continue
            
            prescription_str = str(prescription).strip()
            if not prescription_str:
                continue
            
            # 추가 정보 읽기
            product_name = ws.cell(row_idx, col_map['product_name']).value
            formulation = ws.cell(row_idx, col_map['formulation']).value
            preservative = ws.cell(row_idx, col_map['preservative']).value
            
            # 딕셔너리에 저장
            data_dict[prescription_str] = {
                'product_name': str(product_name).strip() if product_name else '',
                'formulation': str(formulation).strip() if formulation else '',
                'preservative_info': str(preservative).strip() if preservative else ''
            }
            
            data_count += 1
        
        logger.info(f"  ✅ 진행서 데이터 로드 완료: {data_count}개")
        wb.close()
        
        return data_dict
        
    except Exception as e:
        logger.error(f"  ❌ 진행서 읽기 실패: {e}")
        return {}


class PreservationTestOCR:
    """보존력 시험 OCR 전용 클래스"""
    
    def __init__(self, debug_mode=False):
        """
        Azure Document Intelligence 클라이언트 초기화
        
        Args:
            debug_mode: True면 상세 디버깅 정보 출력
        """
        self.endpoint = AZURE_ENDPOINT
        self.key = AZURE_KEY
        self.debug_mode = debug_mode
        
        self.client = DocumentAnalysisClient(
            endpoint=self.endpoint,
            credential=AzureKeyCredential(self.key)
        )
        
        logger.info("✅ Azure Document Intelligence 연결 완료")
        logger.info(f"📍 엔드포인트: {self.endpoint}")
        
        if self.debug_mode:
            logger.info("🐛 디버그 모드 활성화")
    
    def extract_preservation_test_table(self, image_path: str) -> Dict:
        """
        보존력 시험 테이블 추출
        
        Returns:
            {
                'data': [
                    {
                        'test_number': '25E15I14',
                        'prescription_number': 'GB1919-ZMB',
                        'strain': 'E.coli',
                        'cfu_0day': '2.1×10^6',
                        'cfu_7day': '<10^2',
                        'cfu_14day': '<10',
                        'cfu_28day': '<10',
                        'judgment': '적합',
                        'final_judgment': '적합'
                    }
                ],
                'date_info': {
                    'date_0': '01/15',
                    'date_7': '01/22',
                    'date_14': '01/29',
                    'date_28': '02/12'
                }
            }
        """
        logger.info(f"\n🔍 이미지 분석 시작: {os.path.basename(image_path)}")
        
        with open(image_path, 'rb') as f:
            image_data = f.read()
        
        logger.info("📊 테이블 구조 분석 중...")
        poller = self.client.begin_analyze_document("prebuilt-layout", document=image_data)
        result = poller.result()
        
        logger.info(f"📋 감지된 테이블 수: {len(result.tables)}")
        for idx, tbl in enumerate(result.tables):
            logger.info(f"  테이블 {idx}: {tbl.row_count}행 x {tbl.column_count}열")
        
        if not result.tables:
            logger.error("❌ 테이블을 찾을 수 없습니다.")
            return {'data': [], 'date_info': {}}
        
        # 가장 큰 테이블 선택
        table = max(result.tables, key=lambda t: t.row_count * t.column_count)
        logger.info(f"✅ 선택된 테이블: {table.row_count}행 x {table.column_count}열")
        
        # 테이블 매트릭스 생성
        table_matrix = {}
        for cell in table.cells:
            row_idx = cell.row_index
            col_idx = cell.column_index
            if row_idx not in table_matrix:
                table_matrix[row_idx] = {}
            table_matrix[row_idx][col_idx] = cell.content.strip()
        
        # 날짜 정보 추출
        date_info = self._extract_date_info(table_matrix)
        
        # 균주 데이터 추출
        test_data = self._extract_strain_data(table_matrix, date_info)
        
        return {
            'data': test_data,
            'date_info': date_info
        }
    
    def _extract_date_info(self, table_matrix: Dict) -> Dict:
        """
        날짜 정보 추출
        
        처음 5행에서 날짜 패턴 찾기:
        - MM/DD 형식
        - MM DD 형식 (공백 분리)
        - 연속된 4개 날짜
        
        OCR 오인식 대응:
        - '0.5 15' → '05 15' (소수점을 공백으로 인식)
        - '0.6 12' → '06 12'
        """
        logger.info("\n📅 날짜 정보 추출 시도")
        
        date_info = {}
        
        # 처음 5행 확인
        for row_idx in range(min(5, len(table_matrix))):
            if row_idx not in table_matrix:
                continue
            
            row_data = table_matrix[row_idx]
            
            # 🔧 각 셀을 개별적으로 확인 (공백 분리된 날짜 처리)
            dates = []
            for col_idx in sorted(row_data.keys()):
                value = str(row_data[col_idx]).strip()
                
                # 🔧 OCR 오인식 보정: '0.5 15' → '05 15'
                # 패턴: 숫자.숫자 공백 숫자 → 숫자숫자 공백 숫자
                value_corrected = re.sub(r'^(\d)\.(\d)\s+(\d{1,2})$', r'\1\2 \3', value)
                if value_corrected != value:
                    logger.info(f"  🔧 날짜 보정 (Col_{col_idx}): '{value}' → '{value_corrected}'")
                    value = value_corrected
                
                # 패턴 1: MM/DD 또는 MM-DD
                match1 = re.match(r'^(\d{1,2})[/\-.](\d{1,2})$', value)
                if match1:
                    dates.append((match1.group(1), match1.group(2)))
                    continue
                
                # 패턴 2: MM DD (공백으로 분리) - OCR이 자주 이렇게 인식
                match2 = re.match(r'^(\d{1,2})\s+(\d{1,2})$', value)
                if match2:
                    dates.append((match2.group(1), match2.group(2)))
                    logger.info(f"  📍 공백 분리 날짜 발견 (Col_{col_idx}): '{value}' → {match2.group(1)}/{match2.group(2)}")
                    continue
            
            logger.info(f"  행 {row_idx}: 발견된 날짜 {len(dates)}개")
            for i, (m, d) in enumerate(dates[:4]):
                logger.info(f"    날짜 {i}: {m}/{d}")
            
            if len(dates) >= 4:
                # 4개 이상 날짜가 있으면 처음 4개 사용
                date_info = {
                    'date_0': f"{dates[0][0].zfill(2)}/{dates[0][1].zfill(2)}",
                    'date_7': f"{dates[1][0].zfill(2)}/{dates[1][1].zfill(2)}",
                    'date_14': f"{dates[2][0].zfill(2)}/{dates[2][1].zfill(2)}",
                    'date_28': f"{dates[3][0].zfill(2)}/{dates[3][1].zfill(2)}"
                }
                logger.info(f"✅ 날짜 정보 확정: {date_info}")
                return date_info
        
        # 날짜를 찾지 못한 경우 단일 날짜 기반 계산
        logger.info("  ⚠️ 4개 날짜를 찾지 못함. 단일 날짜 기반 계산 시도")
        for row_idx in range(min(5, len(table_matrix))):
            if row_idx not in table_matrix:
                continue
            
            row_data = table_matrix[row_idx]
            
            for col_idx, value in row_data.items():
                value_str = str(value).strip()
                
                # 🔧 OCR 오인식 보정: '0.5 15' → '05 15'
                value_corrected = re.sub(r'^(\d)\.(\d)\s+(\d{1,2})$', r'\1\2 \3', value_str)
                if value_corrected != value_str:
                    logger.info(f"  🔧 날짜 보정: '{value_str}' → '{value_corrected}'")
                    value_str = value_corrected
                
                # 패턴 1: MM/DD 또는 MM-DD
                match1 = re.match(r'^(\d{1,2})[/\-.](\d{1,2})$', value_str)
                if match1:
                    try:
                        month = int(match1.group(1))
                        day = int(match1.group(2))
                        start_date = datetime(2024, month, day)
                        
                        date_info = {
                            'date_0': start_date.strftime("%m/%d"),
                            'date_7': (start_date + timedelta(days=7)).strftime("%m/%d"),
                            'date_14': (start_date + timedelta(days=14)).strftime("%m/%d"),
                            'date_28': (start_date + timedelta(days=28)).strftime("%m/%d")
                        }
                        logger.info(f"✅ 시작 날짜 기반 계산: {date_info}")
                        return date_info
                    except:
                        continue
                
                # 패턴 2: MM DD (공백 분리)
                match2 = re.match(r'^(\d{1,2})\s+(\d{1,2})$', value_str)
                if match2:
                    try:
                        month = int(match2.group(1))
                        day = int(match2.group(2))
                        start_date = datetime(2024, month, day)
                        
                        date_info = {
                            'date_0': start_date.strftime("%m/%d"),
                            'date_7': (start_date + timedelta(days=7)).strftime("%m/%d"),
                            'date_14': (start_date + timedelta(days=14)).strftime("%m/%d"),
                            'date_28': (start_date + timedelta(days=28)).strftime("%m/%d")
                        }
                        logger.info(f"✅ 공백 분리 날짜 기반 계산: {value_str} → {date_info}")
                        return date_info
                    except:
                        continue
        
        logger.warning("⚠️ 날짜 정보 없음")
        return {}
    
    def _extract_strain_data(self, table_matrix: Dict, date_info: Dict) -> List[Dict]:
        """
        균주 데이터 추출
        
        테이블 구조:
        - 시험번호/처방번호: 좌측 상단
        - 균주명: 세로 방향
        - CFU 값: 가로 방향 (0, 7, 14, 28일)
        
        개선: 각 행의 Bulk Name에서 직접 시험번호 추출하여 행 순서 유지
        """
        logger.info("\n🦠 균주 데이터 추출 시작")
        
        # 헤더 행 찾기 (균주명이 있는 행)
        header_row = self._find_header_row(table_matrix)
        
        if header_row is None:
            logger.warning("⚠️ 헤더 행을 찾을 수 없습니다")
            return []
        
        # 🔧 헤더가 없는 경우 (header_row = -1)
        if header_row == -1:
            logger.info("📍 헤더 없음 → 데이터 행부터 시작 (행 0)")
            
            # 첫 번째 데이터 행을 기준으로 컬럼 매핑
            column_map = self._identify_columns(table_matrix, 0)
            logger.info(f"📋 컬럼 매핑: {column_map}")
            
            # 디버그 모드
            if self.debug_mode:
                self._debug_table_structure(table_matrix, 0, column_map)
            
            # 균주 데이터 추출 (행 0부터)
            data_start_row = 0
        else:
            logger.info(f"📍 헤더 행: {header_row}")
            
            # 컬럼 매핑 찾기
            column_map = self._identify_columns(table_matrix, header_row)
            logger.info(f"📋 컬럼 매핑: {column_map}")
            
            # 디버그 모드
            if self.debug_mode:
                self._debug_table_structure(table_matrix, header_row, column_map)
            
            # 균주 데이터 추출 (헤더 다음 행부터)
            data_start_row = header_row + 1
        
        # 균주 데이터 리스트 초기화
        test_data = []
        
        # 🆕 현재 제품 정보 (각 행에서 직접 추출)
        current_test = ''
        current_prescription = ''
        
        # 🆕 Bulk Name 컬럼 찾기 (Col_0)
        bulk_name_col = 0  # 일반적으로 첫 번째 컬럼
        
        for row_idx in range(data_start_row, len(table_matrix)):
            if row_idx not in table_matrix:
                continue
            
            row_data = table_matrix[row_idx]
            
            # 🆕 1. Bulk Name 확인 (새 제품인지?)
            bulk_name = row_data.get(bulk_name_col, '').strip()
            if bulk_name:
                # Bulk Name이 있으면 → 새 제품 시작!
                # 이 행에서 직접 시험번호/처방번호 추출
                test_num, presc_num = self._extract_test_info_from_row(bulk_name)
                if test_num:
                    current_test = test_num
                    logger.info(f"  📦 새 제품 시작 (행{row_idx}): 시험번호={test_num}")
                if presc_num:
                    current_prescription = presc_num
                    logger.info(f"  📦 새 제품 시작 (행{row_idx}): 처방번호={presc_num}")
            
            # 2. 균주명 추출
            strain_col = column_map.get('strain_col')
            if strain_col is None or strain_col not in row_data:
                continue
            
            strain = row_data[strain_col].strip()
            
            # 균주 정규화
            strain_normalized = self._normalize_strain_name(strain)
            if not strain_normalized:
                continue
            
            # 3. CFU 값 추출
            cfu_0 = self._clean_cfu_value(row_data.get(column_map.get('cfu_0_col', -1), ''), strain_normalized, '0일')
            cfu_7 = self._clean_cfu_value(row_data.get(column_map.get('cfu_7_col', -1), ''), strain_normalized, '7일')
            cfu_14 = self._clean_cfu_value(row_data.get(column_map.get('cfu_14_col', -1), ''), strain_normalized, '14일')
            cfu_28 = self._clean_cfu_value(row_data.get(column_map.get('cfu_28_col', -1), ''), strain_normalized, '28일')
            
            # 4. 판정 추출
            judgment = self._extract_judgment(row_data.get(column_map.get('judgment_col', -1), ''))
            
            # 🔧 최종판정: 컬럼이 없거나 값이 없으면 빈 문자열
            final_judgment_col = column_map.get('final_judgment_col', -1)
            if final_judgment_col == -1:
                # 최종판정 컬럼이 없으면 빈 문자열
                final_judgment = ''
            else:
                final_judgment_value = row_data.get(final_judgment_col, '')
                if final_judgment_value:
                    final_judgment = self._extract_judgment(final_judgment_value)
                else:
                    # 값이 비어있으면 빈 문자열
                    final_judgment = ''
            
            test_data.append({
                'test_number': current_test,
                'prescription_number': current_prescription,
                'strain': strain_normalized,
                'cfu_0day': cfu_0,
                'cfu_7day': cfu_7,
                'cfu_14day': cfu_14,
                'cfu_28day': cfu_28,
                'judgment': judgment,
                'final_judgment': final_judgment
            })
        
        logger.info(f"✅ 추출된 균주 데이터: {len(test_data)}개")
        
        # 보정 요약 (간단히)
        logger.info("\n" + "="*80)
        logger.info("📊 데이터 보정 완료")
        logger.info("="*80)
        
        # 균주 순서 정렬 (표준 순서로)
        strain_order = {
            'E.coli': 0,
            'P.aeruginosa': 1,
            'S.aureus': 2,
            'C.albicans': 3,
            'A.brasiliensis': 4
        }
        
        # 시험번호별로 그룹핑하여 정렬
        sorted_data = []
        current_test_number = None
        test_group = []
        
        for data in test_data:
            if data['test_number'] != current_test_number:
                # 이전 그룹 정렬하여 추가
                if test_group:
                    test_group.sort(key=lambda x: strain_order.get(x['strain'], 999))
                    sorted_data.extend(test_group)
                
                # 새 그룹 시작
                current_test_number = data['test_number']
                test_group = [data]
            else:
                test_group.append(data)
        
        # 마지막 그룹 추가
        if test_group:
            test_group.sort(key=lambda x: strain_order.get(x['strain'], 999))
            sorted_data.extend(test_group)
        
        logger.info(f"✅ 균주 순서 정렬 완료: {len(sorted_data)}개")
        
        return sorted_data
    
    def _debug_table_structure(self, table_matrix: Dict, header_row: int, column_map: Dict):
        """
        테이블 구조 상세 디버깅 정보 출력
        
        Args:
            table_matrix: 테이블 매트릭스
            header_row: 헤더 행 인덱스
            column_map: 컬럼 매핑 정보
        """
        logger.info("\n" + "="*80)
        logger.info("🐛 디버그: 테이블 구조 상세 분석")
        logger.info("="*80)
        
        # 1. 테이블 매트릭스 (전체)
        logger.info("\n📋 1. 테이블 매트릭스 (전체)")
        logger.info("-"*80)
        
        for row_idx in range(len(table_matrix)):
            if row_idx in table_matrix:
                logger.info(f"\n행 {row_idx}:")
                row_data = table_matrix[row_idx]
                
                for col_idx in sorted(row_data.keys()):
                    value = row_data[col_idx]
                    display_value = value[:40] if len(value) > 40 else value
                    logger.info(f"  Col_{col_idx}: '{display_value}'")
        
        # 2. 컬럼 매핑 상세
        logger.info("\n" + "="*80)
        logger.info("📊 2. 컬럼 매핑 상세")
        logger.info("-"*80)
        
        for col_name, col_idx in sorted(column_map.items(), key=lambda x: x[1]):
            logger.info(f"  {col_name:20s} → Col_{col_idx}")
        
        # 3. Bulk Name 컬럼 찾기
        logger.info("\n" + "="*80)
        logger.info("🔍 3. Bulk Name 컬럼 찾기")
        logger.info("-"*80)
        
        bulk_name_col = None
        
        # 헤더에서 찾기
        if header_row in table_matrix:
            for col_idx, value in table_matrix[header_row].items():
                if 'BULK' in value.upper() or 'NAME' in value.upper():
                    bulk_name_col = col_idx
                    logger.info(f"✅ Bulk Name 컬럼 발견: Col_{col_idx}")
                    logger.info(f"   헤더 값: '{value}'")
                    break
        
        if bulk_name_col is None:
            logger.info("⚠️ Bulk Name 컬럼을 찾을 수 없습니다. Col_0으로 가정")
            bulk_name_col = 0
        
        # 4. Bulk Name 컬럼 전체 내용
        logger.info("\n" + "="*80)
        logger.info(f"📝 4. Bulk Name 컬럼(Col_{bulk_name_col}) 전체 내용")
        logger.info("-"*80)
        
        for row_idx in sorted(table_matrix.keys()):
            if bulk_name_col in table_matrix[row_idx]:
                value = table_matrix[row_idx][bulk_name_col]
                if value:  # 빈 값 제외
                    logger.info(f"행 {row_idx:2d}: '{value}'")
        
        # 5. 시험번호/처방번호 패턴 찾기
        logger.info("\n" + "="*80)
        logger.info("🔬 5. 시험번호/처방번호 패턴 매칭")
        logger.info("-"*80)
        
        # 시험번호 패턴
        test_pattern = r'\b(\d{2}[A-Z]\d{2}[A-Z]\d{2,3})\b'
        
        # 처방번호 패턴
        presc_patterns = [
            r'\b([A-Z]{2,4}\d{4,5}[A-Z]?-[A-Z]{1,4}\d?)\b',
            r'\b([A-Z]{2,4}\d{4}-[A-Z]{1,5})\b'
        ]
        
        # Bulk Name 컬럼에서 패턴 찾기
        found_tests = []
        found_prescriptions = []
        
        for row_idx in sorted(table_matrix.keys()):
            if bulk_name_col in table_matrix[row_idx]:
                value = table_matrix[row_idx][bulk_name_col]
                
                # 시험번호 찾기
                test_match = re.search(test_pattern, value.upper())
                if test_match:
                    test_num = test_match.group(1)
                    found_tests.append((row_idx, test_num))
                    logger.info(f"✅ 행 {row_idx}: 시험번호 '{test_num}'")
                
                # 처방번호 찾기
                for pattern in presc_patterns:
                    presc_match = re.search(pattern, value.upper())
                    if presc_match:
                        presc_num = presc_match.group(1)
                        found_prescriptions.append((row_idx, presc_num))
                        logger.info(f"✅ 행 {row_idx}: 처방번호 '{presc_num}'")
                        break
        
        # 6. 전체 테이블에서도 검색
        logger.info("\n" + "="*80)
        logger.info("🔍 6. 전체 테이블에서 패턴 검색")
        logger.info("-"*80)
        
        all_text = ' '.join([
            ' '.join(str(v) for v in row.values())
            for row in table_matrix.values()
        ])
        
        test_matches = re.findall(test_pattern, all_text.upper())
        if test_matches:
            logger.info(f"✅ 전체 테이블에서 시험번호 발견: {test_matches}")
        else:
            logger.info("❌ 시험번호 패턴 매칭 실패")
        
        presc_matches = []
        for pattern in presc_patterns:
            matches = re.findall(pattern, all_text.upper())
            presc_matches.extend(matches)
        
        if presc_matches:
            logger.info(f"✅ 전체 테이블에서 처방번호 발견: {presc_matches}")
        else:
            logger.info("❌ 처방번호 패턴 매칭 실패")
        
        # 7. 결론 및 권장사항
        logger.info("\n" + "="*80)
        logger.info("💡 7. 결론 및 권장사항")
        logger.info("-"*80)
        
        if found_tests:
            logger.info(f"\n✅ Bulk Name 컬럼에서 시험번호 {len(found_tests)}개 발견:")
            for row, test in found_tests:
                logger.info(f"   행 {row}: {test}")
            logger.info(f"\n📝 권장: _extract_test_info()를 수정하여 데이터 행도 스캔")
        else:
            logger.info("\n❌ Bulk Name 컬럼에서 시험번호를 찾을 수 없습니다!")
            if test_matches:
                logger.info(f"   하지만 전체 테이블에는 존재: {test_matches}")
                logger.info(f"   → 다른 컬럼에 있을 수 있음")
        
        if found_prescriptions:
            logger.info(f"\n✅ Bulk Name 컬럼에서 처방번호 {len(found_prescriptions)}개 발견:")
            for row, presc in found_prescriptions:
                logger.info(f"   행 {row}: {presc}")
        else:
            logger.info("\n⚠️ Bulk Name 컬럼에서 처방번호를 찾을 수 없습니다!")
        
        logger.info("\n" + "="*80)
    
    def _find_header_row(self, table_matrix: Dict) -> Optional[int]:
        """
        헤더 행 찾기
        
        우선순위:
        1. 'Challenged Organism' 또는 'Bulk Name' 텍스트가 있는 행
        2. 균주명 키워드가 있는 행
        """
        # 1순위: 명확한 헤더 키워드로 찾기
        header_keywords = ['CHALLENGED ORGANISM', 'BULK NAME', 'SPECIFICATION']
        
        for row_idx in range(min(5, len(table_matrix))):
            if row_idx not in table_matrix:
                continue
            
            row_text = ' '.join(str(v) for v in table_matrix[row_idx].values()).upper()
            
            # 헤더 키워드가 있으면 진짜 헤더
            if any(keyword in row_text for keyword in header_keywords):
                logger.info(f"  ✅ 헤더 행 발견 (명확한 키워드): 행 {row_idx}")
                return row_idx
        
        # 2순위: 균주명 키워드로 찾기 (데이터 행과 구분 필요!)
        strain_keywords = ['E.COLI', 'ESCHERICHIA', 'P.AERUGINOSA', 'PSEUDOMONAS', 
                          'S.AUREUS', 'STAPHYLOCOCCUS', 'C.ALBICANS', 'CANDIDA',
                          'A.BRASILIENSIS', 'ASPERGILLUS', '균주', 'STRAIN']
        
        candidate_rows = []
        
        for row_idx in range(min(15, len(table_matrix))):
            if row_idx not in table_matrix:
                continue
            
            row_text = ' '.join(str(v) for v in table_matrix[row_idx].values()).upper()
            
            if any(keyword in row_text for keyword in strain_keywords):
                candidate_rows.append(row_idx)
        
        # 후보 행이 있으면 진짜 헤더인지 데이터 행인지 판별
        for row_idx in candidate_rows:
            row_data = table_matrix[row_idx]
            
            # 🔍 헤더 행 판별: CFU 값 패턴이 없어야 함
            has_cfu_pattern = False
            
            for col_idx, value in row_data.items():
                value_str = str(value).strip()
                
                # CFU 값 패턴 (과학적 표기법)
                if re.search(r'\d+\.?\d*\s*[×xX]\s*10[\^]?\d+', value_str):
                    has_cfu_pattern = True
                    break
                
                # 숫자 패턴 (≤3 같은 Specification 제외)
                if re.match(r'^\d{4,}$', value_str):  # 4자리 이상 숫자
                    has_cfu_pattern = True
                    break
            
            if has_cfu_pattern:
                # CFU 값이 있으면 데이터 행! 헤더가 없는 경우
                logger.info(f"  🔍 행 {row_idx}는 데이터 행 (헤더 없음)")
                logger.info(f"  ✅ 헤더 없이 데이터부터 시작 → header_row = -1")
                return -1  # 특수 값: 헤더 없음
            else:
                # CFU 값이 없으면 진짜 헤더
                logger.info(f"  ✅ 헤더 행 발견 (균주 키워드): 행 {row_idx}")
                return row_idx
        
        logger.warning("  ❌ 헤더 행을 찾을 수 없습니다")
        return None
    
    def _identify_columns(self, table_matrix: Dict, header_row: int) -> Dict:
        """
        컬럼 매핑 생성
        
        Returns:
            {
                'strain_col': int,      # 균주명 컬럼
                'specification_col': int,  # Specification 컬럼 (건너뛰기)
                'cfu_0_col': int,       # 0일 CFU 컬럼
                'cfu_7_col': int,       # 7일 CFU 컬럼
                'cfu_14_col': int,      # 14일 CFU 컬럼
                'cfu_28_col': int,      # 28일 CFU 컬럼
                'judgment_col': int,    # 판정 컬럼
                'final_judgment_col': int  # 최종판정 컬럼
            }
        """
        column_map = {}
        
        if header_row not in table_matrix:
            return column_map
        
        header_data = table_matrix[header_row]
        
        # 균주명 컬럼 찾기
        for col_idx, value in header_data.items():
            value_upper = value.upper().strip()
            
            # 균주 컬럼
            if '균주' in value or 'STRAIN' in value_upper or 'E.COLI' in value_upper or 'ORGANISM' in value_upper:
                column_map['strain_col'] = col_idx
                logger.info(f"  ✅ 균주 컬럼 감지: Col_{col_idx}")
            
            # ⭐ Specification 컬럼 (건너뛰어야 함) - 개선된 감지
            if 'SPECIFICATION' in value_upper or 'SPEC' in value_upper:
                column_map['specification_col'] = col_idx
                logger.info(f"  ⚠️ Specification 컬럼 감지: Col_{col_idx} (건너뜀)")
            
            # CFU 컬럼 (0, 7, 14, 28 숫자 찾기)
            if '0' in value and ('일' in value or 'DAY' in value_upper or 'CFU' in value_upper or '접종' in value):
                column_map['cfu_0_col'] = col_idx
            elif '7' in value and ('일' in value or 'DAY' in value_upper or 'CFU' in value_upper):
                column_map['cfu_7_col'] = col_idx
            elif '14' in value and ('일' in value or 'DAY' in value_upper or 'CFU' in value_upper):
                column_map['cfu_14_col'] = col_idx
            elif '28' in value and ('일' in value or 'DAY' in value_upper or 'CFU' in value_upper):
                column_map['cfu_28_col'] = col_idx
            
            # 판정 컬럼
            if '판정' in value or 'JUDGMENT' in value_upper:
                if '최종' in value or 'FINAL' in value_upper:
                    column_map['final_judgment_col'] = col_idx
                elif 'judgment_col' not in column_map:
                    column_map['judgment_col'] = col_idx
        
        # ⭐ 추론: CFU 컬럼이 없으면 균주 컬럼 다음부터 순서대로 할당
        # 단, Specification 컬럼은 건너뛰기!
        if 'strain_col' in column_map:
            strain_col = column_map['strain_col']
            spec_col = column_map.get('specification_col', -1)
            
            # 🔧 헤더가 없을 때: Specification 컬럼을 값 패턴으로 감지
            if spec_col == -1:
                # 균주 다음 컬럼이 Specification일 가능성 체크
                next_col = strain_col + 1
                
                # 여러 행 확인 (최소 3개 행)
                spec_pattern_count = 0
                checked_rows = 0
                
                for row_idx in sorted(table_matrix.keys()):
                    if row_idx <= header_row:
                        continue  # 헤더 행은 건너뛰기
                    
                    if checked_rows >= 5:  # 5개 행만 확인
                        break
                    
                    if next_col in table_matrix[row_idx]:
                        value = str(table_matrix[row_idx][next_col]).strip()
                        
                        # Specification 값 패턴
                        # ≤3, ≤1, ≤0, ≤0°, 53, 51, 50, 50c 등
                        if re.match(r'^(≤[0-9]+[°cC]?|[0-9]{1,2}[°cC]?|SI)$', value):
                            spec_pattern_count += 1
                        
                        checked_rows += 1
                
                # 3개 이상 행에서 Specification 패턴 발견하면 감지
                if spec_pattern_count >= 3:
                    spec_col = next_col
                    column_map['specification_col'] = spec_col
                    logger.info(f"  🔧 Specification 컬럼 추론: Col_{spec_col} (값 패턴 기반)")
            
            # CFU 시작 컬럼 계산
            if spec_col > strain_col:
                # Specification이 균주 다음에 있으면 그 다음부터
                cfu_start_col = spec_col + 1
                logger.info(f"  📍 CFU 시작 위치: Col_{cfu_start_col} (Specification 건너뜀)")
            else:
                # Specification이 없거나 균주 앞에 있으면 균주 다음부터
                cfu_start_col = strain_col + 1
                logger.info(f"  📍 CFU 시작 위치: Col_{cfu_start_col}")
            
            if 'cfu_0_col' not in column_map:
                column_map['cfu_0_col'] = cfu_start_col
            if 'cfu_7_col' not in column_map:
                column_map['cfu_7_col'] = cfu_start_col + 1
            if 'cfu_14_col' not in column_map:
                column_map['cfu_14_col'] = cfu_start_col + 2
            if 'cfu_28_col' not in column_map:
                column_map['cfu_28_col'] = cfu_start_col + 3
            if 'judgment_col' not in column_map:
                column_map['judgment_col'] = cfu_start_col + 4
            if 'final_judgment_col' not in column_map:
                column_map['final_judgment_col'] = cfu_start_col + 5
        
        return column_map
    
    def _extract_test_info_from_row(self, row_text: str) -> Tuple[str, str]:
        """
        단일 행 텍스트에서 시험번호와 처방번호 추출
        
        Args:
            row_text: Bulk Name 컬럼의 텍스트
            
        Returns:
            (시험번호, 처방번호) 튜플
        """
        import re
        
        test_number = ''
        prescription_number = ''
        
        if not row_text:
            return '', ''
        
        # 전처리
        row_text_upper = row_text.upper()
        row_text_upper = row_text_upper.replace('!', 'I')  # OCR 오류 보정
        row_text_upper = row_text_upper.replace('|', 'I')  # | → I 보정
        row_text_upper = re.sub(r'-\s+', '-', row_text_upper)  # '- ' → '-'
        row_text_upper = re.sub(r'\s+-', '-', row_text_upper)  # ' -' → '-'
        row_text_upper = re.sub(r'-+', '-', row_text_upper)   # '--', '---' → '-'
        row_text_upper = re.sub(r'\s+', ' ', row_text_upper)   # 연속 공백 → 단일
        
        # 1. 시험번호 패턴 (25A15I14, 25E15114 등)
        test_patterns = [
            r'\b(2[0-9][A-Z]\d{2}[I!|1]\d{2})\b',  # 25A15I14, 25A15|14
            r'\b(2[0-9][E]\d{2}1\d{2})\b',         # 25E15114 (I → 1)
        ]
        
        for pattern in test_patterns:
            match = re.search(pattern, row_text_upper)
            if match:
                test_number = match.group(1)
                # 보정: 1 → I
                test_number = re.sub(r'([A-Z])(\d{2})1(\d{2})', r'\g<1>\g<2>I\g<3>', test_number)
                # 보정: | → I
                test_number = test_number.replace('|', 'I')
                # 보정: ! → I
                test_number = test_number.replace('!', 'I')
                break
        
        # 2. 처방번호 패턴
        prescription_patterns = [
            # 기본 패턴
            r'\b([A-Z]{2,4}\d{4,5}[A-Z]?-[A-Z]{1,5}\d?)\b',
            r'\b([A-Z]{3}\d{5}-[A-Z]{2,4})\b',
            r'\b(M-[A-Z]{2,4}\d{4,5}-[A-Z]{1,4}\d?)\b',
            
            # 확장 패턴
            r'\b([A-Z]{2,4}\d{3,6}-[A-Z]{1,5})\b',
            r'\b([A-Z]{2,5}\d{4}-[A-Z]{1,3}\d{0,2})\b',
            r'\b([A-Z]{1,3}\d{4,5}-[A-Z]{2,4}[A-Z]?)\b',
            r'\b([A-Z]{2,4}\d{4}-[A-Z]\d[A-Z]{1,3})\b',
            r'\b([A-Z]{2,4}\d{3,4}[A-Z]?-[A-Z]{1,4}\d*)\b',
            
            # 🆕 숫자+문자 조합 (11F, 01GC 등)
            r'\b([A-Z]{2,4}\d{4}-\d{1,2}[A-Z]{1,2})\b',
            
            # 공백 허용
            r'\b([A-Z]{2,4}\d{4,5}[A-Z]?-\s*[A-Z]{1,5}\d?)\b',
            
            # 복잡한 접미사
            r'\b([A-Z]{2,4}\d{4,5}[A-Z]?-\s*[A-Z]+\d+[A-Z]+)\b',
            r'\b([A-Z]{2,4}\d{4,5}[A-Z]?-[A-Z]{1,5}\d[A-Z]+)\b',
            
            # AZLY1 타입
            r'\b([A-Z]{2,4}\d{3,5}-[A-Z]{1,4}\d{1,2})\b',
            
            # 포괄적 패턴
            r'\b([A-Z]{2,5}\d{3,5}-[A-Z]{2,5}[A-Z\d]*)\b',
        ]
        
        for pattern in prescription_patterns:
            match = re.search(pattern, row_text_upper)
            if match:
                prescription_number = match.group(1).strip()
                break
        
        return test_number, prescription_number
    
    def _extract_test_info(self, table_matrix: Dict, header_row: int) -> Tuple[List[str], List[str]]:
        """
        시험번호와 처방번호 추출
        
        전체 행을 스캔하여 Bulk Name 컬럼에서 패턴 찾기
        OCR 오인식 대응: 25E15I14 → 25E15114 (I가 1로 인식됨)
        """
        test_numbers = []
        prescription_numbers = []
        
        # 🔧 전체 행 스캔
        for row_idx in range(len(table_matrix)):
            if row_idx not in table_matrix:
                continue
            
            row_data = table_matrix[row_idx]
            row_text = ' '.join(str(v) for v in row_data.values())
            
            # 전처리
            row_text_upper = row_text.upper()
            row_text_upper = row_text_upper.replace('!', 'I')  # OCR 오류 보정
            row_text_upper = row_text_upper.replace('|', 'I')  # | → I 보정
            row_text_upper = re.sub(r'-\s+', '-', row_text_upper)  # '- ' → '-'
            row_text_upper = re.sub(r'\s+-', '-', row_text_upper)  # ' -' → '-'
            row_text_upper = re.sub(r'-+', '-', row_text_upper)   # '--', '---' → '-'
            row_text_upper = re.sub(r'\s+', ' ', row_text_upper)   # 연속 공백 → 단일
            
            # ========================================
            # 1. 처방번호 추출 (고도화된 패턴)
            # ========================================
            prescription_patterns = [
                # 기본 패턴
                r'\b[A-Z]{2,4}\d{4,5}[A-Z]?-[A-Z]{1,5}\d?\b',
                r'\b[A-Z]{3}\d{5}-[A-Z]{2,4}\b',
                r'\bM-[A-Z]{2,4}\d{4,5}-[A-Z]{1,4}\d?\b',
                
                # 확장 패턴
                r'\b[A-Z]{2,4}\d{3,6}-[A-Z]{1,5}\b',
                r'\b[A-Z]{2,5}\d{4}-[A-Z]{1,3}\d{0,2}\b',
                r'\b[A-Z]{1,3}\d{4,5}-[A-Z]{2,4}[A-Z]?\b',
                r'\b[A-Z]{2,4}\d{4}-[A-Z]\d[A-Z]{1,3}\b',
                r'\b[A-Z]{2,4}\d{3,4}[A-Z]?-[A-Z]{1,4}\d*\b',
                
                # 🆕 숫자+문자 조합 (11F, 01GC 등)
                r'\b[A-Z]{2,4}\d{4}-\d{1,2}[A-Z]{1,2}\b',  # WC1820-11F
                
                # 공백 허용
                r'\b[A-Z]{2,4}\d{4,5}[A-Z]?-\s*[A-Z]{1,5}\d?\b',
                
                # 🆕 복잡한 접미사 (RZ9A, OZ2A 등 - 중간 숫자)
                r'\b[A-Z]{2,4}\d{4,5}[A-Z]?-\s*[A-Z]+\d+[A-Z]+\b',  # RZ9A
                r'\b[A-Z]{2,4}\d{4,5}[A-Z]?-[A-Z]{1,5}\d[A-Z]+\b',  # OZ2A
                
                # AZLY1 타입
                r'\b[A-Z]{2,4}\d{3,5}-[A-Z]{1,4}\d{1,2}\b',
                
                # 포괄적 패턴
                r'\b[A-Z]{2,5}\d{3,5}-[A-Z]{2,5}[A-Z\d]*\b',
            ]
            
            for pattern in prescription_patterns:
                matches = re.findall(pattern, row_text_upper)
                for match in matches:
                    # 공백 제거 정규화
                    normalized = match.replace(' ', '')
                    normalized = re.sub(r'-+', '-', normalized)  # 여러 대시 → 1개
                    
                    if normalized not in prescription_numbers:
                        prescription_numbers.append(normalized)
                        if match != normalized:
                            logger.info(f"  ✅ 처방번호 발견 (행{row_idx}): {match} → {normalized} (정규화)")
                        else:
                            logger.info(f"  ✅ 처방번호 발견 (행{row_idx}): {match}")
            
            # ========================================
            # 2. 시험번호 추출 (고도화된 패턴)
            # ========================================
            
            # 🔧 중복 방지를 위한 임시 리스트
            found_in_this_row = []
            
            # 2-1. 정상 형태 (I가 정확히 인식된 경우)
            correct_matches = re.findall(r'\b(\d{2}[A-L]\d{2}I\d{2,3})\b', row_text_upper)
            for match in correct_matches:
                if match not in test_numbers and match not in found_in_this_row:
                    test_numbers.append(match)
                    found_in_this_row.append(match)
                    logger.info(f"  ✅ 시험번호 발견 (행{row_idx}): {match}")
            
            # 2-2. OCR 오류 형태 (I를 1로 잘못 인식) - 정상 형태가 없을 때만
            if not correct_matches:
                ocr_error_matches = re.findall(r'\b(\d{2}[A-L]\d{2}1\d{2,3})\b', row_text_upper)
                for match in ocr_error_matches:
                    # I/1 보정: 25A15114 → 25A15I14
                    corrected = match[:5] + 'I' + match[6:]
                    if corrected not in test_numbers and corrected not in found_in_this_row:
                        test_numbers.append(corrected)
                        found_in_this_row.append(corrected)
                        logger.info(f"  ✅ 시험번호 발견 (행{row_idx}): {match} → {corrected} (I/1 보정)")
            
            # 2-3. I가 누락된 형태 (숫자만 연속) - 정상 형태와 I/1 보정이 없을 때만
            if not correct_matches and not ocr_error_matches:
                missing_i_matches = re.findall(r'\b(\d{2}[A-L]\d{5,6})\b', row_text_upper)
                for match in missing_i_matches:
                    # I 삽입: 25A15102 → 25A15I02
                    if len(match) == 7:  # 25A2012
                        corrected = match[:5] + 'I' + match[6:]
                    elif len(match) == 8:  # 25A20102
                        corrected = match[:5] + 'I' + match[5:]
                    else:
                        continue
                    
                    if corrected not in test_numbers and corrected not in found_in_this_row:
                        test_numbers.append(corrected)
                        found_in_this_row.append(corrected)
                        logger.info(f"  ✅ 시험번호 발견 (행{row_idx}): {match} → {corrected} (I 삽입)")
            
            # 2-4. 공백이 있는 형태
            space_matches = re.findall(r'(\d{2})([A-L])(\d)\s+(\d)(\d{2,3})', row_text_upper)
            for year_prefix, letter, d1, d2, last_digits in space_matches:
                converted = f"{year_prefix}{letter}{d1}{d2}I{last_digits[:2]}"
                if converted not in test_numbers and converted not in found_in_this_row:
                    test_numbers.append(converted)
                    found_in_this_row.append(converted)
                    logger.info(f"  ✅ 시험번호 발견 (행{row_idx}): {year_prefix}{letter}{d1} {d2}{last_digits} → {converted} (공백 제거)")
            
            # 2-5. 잘린 형태 (줄 끝) - 다른 형태가 없을 때만
            if not found_in_this_row:
                truncated_matches = re.findall(r'(\d{2}[A-L]\d{2}[A-Z1I|]?\d{0,3})\s*$', row_text_upper)
                for match in truncated_matches:
                    # 길이 체크
                    if len(match) < 6:
                        continue
                    
                    # I/1 정규화
                    normalized = match
                    if len(match) >= 6 and match[5] in ['1', '|']:
                        normalized = match[:5] + 'I' + match[6:]
                    
                    # 잘림 경고
                    if len(normalized) < 8:
                        logger.warning(f"  ⚠️ 시험번호 잘림 가능성 (행{row_idx}): {match}")
                    
                    if normalized not in test_numbers:
                        test_numbers.append(normalized)
                        if match != normalized:
                            logger.info(f"  ✅ 시험번호 발견 (행{row_idx}): {match} → {normalized} (정규화)")
                        else:
                            logger.info(f"  ✅ 시험번호 발견 (행{row_idx}): {match}")
        
        logger.info(f"📝 시험번호: {test_numbers}")
        logger.info(f"📝 처방번호: {prescription_numbers}")
        
        # 최소 1개는 있어야 함
        if not test_numbers:
            test_numbers = ['']
        if not prescription_numbers:
            prescription_numbers = ['']
        
        return test_numbers, prescription_numbers
    
    def _normalize_strain_name(self, strain: str) -> str:
        """균주명 정규화"""
        strain_mapping = {
            'E.coli': 'E.coli',
            'Escherichia coli': 'E.coli',
            'E. coli': 'E.coli',
            'Escherichia': 'E.coli',
            
            'P.aeruginosa': 'P.aeruginosa',
            'Pseudomonas aeruginosa': 'P.aeruginosa',
            'P. aeruginosa': 'P.aeruginosa',
            'Pseudomonas': 'P.aeruginosa',
            
            'S.aureus': 'S.aureus',
            'Staphylococcus aureus': 'S.aureus',
            'S. aureus': 'S.aureus',
            'Staphylococcus': 'S.aureus',
            
            'C.albicans': 'C.albicans',
            'Candida albicans': 'C.albicans',
            'C. albicans': 'C.albicans',
            'Candida': 'C.albicans',
            
            'A.brasiliensis': 'A.brasiliensis',
            'Aspergillus brasiliensis': 'A.brasiliensis',
            'A. brasiliensis': 'A.brasiliensis',
            'Aspergillus': 'A.brasiliensis'
        }
        
        for full_name, short_name in strain_mapping.items():
            if full_name.lower() in strain.lower():
                return short_name
        
        return ''
    
    def _split_merged_cells(self, value: str) -> str:
        """
        합쳐진 셀 감지 및 첫 번째 값 추출
        
        OCR이 두 셀을 하나로 합쳐버리는 경우:
        - '<10 < 10"' → 두 개의 <10
        - '7.0X102 1.0 ×103' → 7.0×10^2와 1.0×10^3
        - '< 10" < 10}' → 두 개의 <10
        
        Args:
            value: 원본 값
            
        Returns:
            첫 번째 값만 추출
        """
        if not value:
            return value
        
        import re
        
        # 패턴 1: 과학적 표기법이 2개 이상 (공백으로 구분)
        # '7.0X102 1.0 ×103' 같은 패턴
        scientific_pattern = r'(\d+\.?\d*[×xX]10[\^]?\d+)'
        matches = re.findall(scientific_pattern, value)
        
        if len(matches) >= 2:
            logger.warning(f"  ⚠️ 합쳐진 셀 감지: '{value}' → 첫 번째 값만 사용: '{matches[0]}'")
            return matches[0]
        
        # 패턴 2: < 기호가 2개 이상
        # '<10 < 10"' 같은 패턴
        less_than_pattern = r'<\s*\d+'
        less_matches = re.findall(less_than_pattern, value)
        
        if len(less_matches) >= 2:
            logger.warning(f"  ⚠️ 합쳐진 셀 감지: '{value}' → 첫 번째 값만 사용: '{less_matches[0]}'")
            return less_matches[0]
        
        return value
    
    def _remove_noise(self, value: str) -> str:
        """
        불필요한 노이즈 제거
        
        Args:
            value: 원본 값
            
        Returns:
            노이즈가 제거된 값
        """
        if not value:
            return value
        
        # :selected:, :unselected: 제거
        value = value.replace(':selected:', '').replace(':unselected:', '')
        
        # 따옴표 제거
        value = value.replace('"', '').replace("'", '')
        
        # 도 기호 제거
        value = value.replace('°', '')
        
        # 유로 기호 제거
        value = value.replace('€', '')
        
        # 줄바꿈 제거
        value = value.replace('\n', ' ')
        
        # 앞뒤 공백 제거
        value = value.strip()
        
        return value
    
    def _fix_less_than_10(self, value: str) -> str:
        """
        <10의 명확한 오인식 패턴 수정
        
        주의: <10^2 (100 미만)와 <10 (10 미만)은 다름!
        
        OCR 오인식 패턴:
        - 40, 40°, 40€  → <10
        - CIO, CIÒ      → <10
        - C10, 410      → <10
        - < 10" (따옴표) → <10
        - < 10 2        → <10^2 (지수!)
        - <102, < 102   → <10^2 (지수!)
        - 4102          → <10^2 (< → 4 오인식)
        - <12, <62      → <10^2 (0 누락)
        - 110, 210      → <10 (< 누락)
        - <1>, LU       → <10 (심각한 오인식)
        
        Args:
            value: 원본 값
            
        Returns:
            수정된 값
        """
        if not value:
            return value
        
        value = value.strip()
        
        # 🆕 0. 의미 없는 값 → 빈 문자열
        if value in ['...', '....', '…']:
            return ''
        
        # 1. 명확한 <10 오인식 패턴 (숫자가 틀린 경우)
        # 기존 + 추가
        less_than_10_patterns = [
            '40', '40°', '40€',  # 기존
            'CIO', 'CIÒ', 'C10', '410', '90',  # 기존
            # 🆕 추가 패턴
            'Lio', 'LIO', 'Clo', 'CLO',  # <10 → Lio/Clo
            'CO', 'cio', 'clo',  # 소문자 버전
            'L10', 'L 10', 'L10"', 'L 10"',  # < → L
            '€10', '€ 10',  # < → €
            '010', '(10)', '(10', '10)',  # < → 0/(
            '(1)', '(1', '1)',  # <10 → (1)
            '2 <10',  # 노이즈 + <10
            # 🆕 2차 추가 (LION, zion 계열)
            'LION', 'LION,', 'Lion', 'lion',  # < → L, 0 → O
            'zion', 'Zion', 'ZION',  # < → z, 0 → o
            # 🆕 숫자 뒤에 L
            '40L', '10L',
            # 🆕 뒤에 0 추가
            '400', '4100',
            # 🆕 6 → <
            '610',
            # 🆕 3차 추가 (새로 발견)
            'Cle', 'CLE', 'Cia', 'CIA',  # <10 → Cle/Cia
            'CCO', 'cco',  # <10 → CCO
            '00',  # <10 → 00
            'COL', 'Col',  # <10 → COL
            'clo"', 'clo\'',  # 따옴표 포함
        ]
        
        if value in less_than_10_patterns:
            return '<10'
        
        import re
        
        # 🆕 1-1. 특수문자 포함 패턴
        # <10?, <10-, <10) 등
        if re.match(r'^<\s*10[\?\-\)]+$', value):
            return '<10'
        
        # 🆕 1-2. < cion 같은 복잡한 패턴
        if re.match(r'^<\s*[czsCZS]ion', value, re.IGNORECASE):
            return '<10'
        
        # 🆕 1-3. 단일 숫자 (1, 2) → <10 (손글씨에서 < 완전 누락)
        if re.match(r'^\d$', value):  # 한 자리 숫자
            return '<10'
        
        # 🆕 1-4. 두 자리 숫자 (00) → <10
        if value == '00':
            return '<10'
        
        # 2. <10^2 패턴들
        
        # 2-1. 정상 패턴: <102, < 102, <10^2
        if re.match(r'^<\s*10[\^]?2$', value):
            return '<10^2'
        
        # 🆕 2-1-1. 쉼표 포함: <102,
        if re.match(r'^<\s*10[\^]?2,?$', value):
            return '<10^2'
        
        # 2-2. < 10 2 (공백으로 분리)
        if re.match(r'^<\s*10\s+2$', value):
            return '<10^2'
        
        # 2-3. 4102, 5102 (< → 4/5 오인식, 7일차에서 빈번)
        if value in ['4102', '5102', '6102', '512']:
            return '<10^2'
        
        # 2-4. <12, <62 (0 누락)
        if value in ['<12', '<62', '<1.2']:
            return '<10^2'
        
        # 🆕 2-5. GIO2, CIS2, C12 (< → G/C 오인식)
        if value in ['GIO2', 'GI02', 'CIS2', 'C12', 'C102']:
            return '<10^2'
        
        # 🆕 2-6. CIO2, Clo2 (< → C 오인식, 1 → I/l)
        if value in ['CIO2', 'Clo2', 'CI02', 'ClO2']:
            return '<10^2'
        
        # 🆕 2-6-1. 쉼표 포함: SI02,
        if re.match(r'^[SC]I0?2,?$', value, re.IGNORECASE):
            return '<10^2'
        
        # 🆕 2-7. 5/02, C/02 (< → 숫자/문자/ 오인식)
        if re.match(r'^[5C6]/0?2$', value):  # 5/02, 5/2, C/02
            return '<10^2'
        
        # 🆕 2-8. ( 102, (12 (< → ( 오인식)
        if re.match(r'^\(\s*10?2,?$', value):  # ( 102, (102, (12
            return '<10^2'
        
        # 🆕 2-9. SI02 2, (102 2 (뒤에 노이즈 2)
        if re.match(r'^[SC]I0?2\s+2$', value, re.IGNORECASE):  # SI02 2, (102 2
            return '<10^2'
        
        # 🆕 2-10. 45102 (앞에 노이즈)
        if re.match(r'^\d+[45]102$', value):  # 45102, 34102
            return '<10^2'
        
        # 3. 특수 <10 패턴들
        
        # 3-1. 110, 210, 2103 (< 누락, 14일차에서 빈번)
        if value in ['110', '210', '2103', '510']:
            return '<10'
        
        # 3-2. <1>, LU, /10 (심각한 오인식)
        if value in ['<1>', 'LU', '/10']:
            return '<10'
        
        # 🆕 3-3. 2 <10 같은 노이즈 제거
        if re.match(r'^\d+\s*<\s*10', value):
            return '<10'
        
        # 🆕 4. <10^3 패턴 (새로 추가!)
        # 103 → <10^3
        if value == '103':
            return '<10^3'
        
        # 5. <10 + 따옴표/특수문자만 (숫자 없음)
        if re.match(r'^<\s*10\s*["\'\s\?\-\)]*$', value):
            return '<10'
        
        # 6. 이미 올바른 형태
        if value == '<10' or value == '< 10':
            return '<10'
        
        return value
    
    def _normalize_scientific(self, value: str) -> str:
        """
        과학적 표기법 정규화
        
        변환:
        - 5.5X105    → 5.5×10^5
        - 6.8×105    → 6.8×10^5
        - 6.0 × 10   → 6.0×10^1
        - <6.1 × 100 → <6.1×10^0
        
        Args:
            value: 원본 값
            
        Returns:
            정규화된 값
        """
        if not value:
            return value
        
        value = value.strip()
        
        # X를 ×로 통일
        value = value.replace('X', '×').replace('x', '×')
        
        import re
        
        # 패턴 1: 숫자.숫자 × 10 숫자 (띄어쓰기 있음)
        pattern1 = r'(\d+\.?\d*)\s*[×]\s*10\s*(\d*)'
        match1 = re.search(pattern1, value)
        
        if match1:
            base = match1.group(1)
            exponent = match1.group(2) if match1.group(2) else '0'
            
            # 부등호 유지
            prefix = ''
            if value.startswith('<'):
                prefix = '<'
            elif value.startswith('≤'):
                prefix = '≤'
            
            return f'{prefix}{base}×10^{exponent}'
        
        # 패턴 2: 숫자.숫자×10숫자 (띄어쓰기 없음)
        pattern2 = r'(\d+\.?\d*)[×]10(\d+)'
        match2 = re.search(pattern2, value)
        
        if match2:
            base = match2.group(1)
            exponent = match2.group(2)
            
            prefix = ''
            if value.startswith('<'):
                prefix = '<'
            elif value.startswith('≤'):
                prefix = '≤'
            
            return f'{prefix}{base}×10^{exponent}'
        
        return value
    
    def _clean_cfu_value(self, value: str, strain: str, day_column: str) -> str:
        """
        CFU 값 보정 (통합 파이프라인)
        
        Args:
            value: 원본 값
            strain: 균주명
            day_column: 날짜 컬럼 ('0일', '7일', '14일', '28일')
            
        Returns:
            보정된 값
        """
        if not value:
            return ""
        
        original_value = value
        
        # 0단계: 합쳐진 셀 분리 (가장 먼저!)
        value = self._split_merged_cells(value)
        
        # 1단계: 노이즈 제거
        value = self._remove_noise(value)
        
        # 2단계: 컬럼별 보정
        if day_column == '0일':
            # 0일: 과학적 표기법만 정규화
            value = self._normalize_scientific(value)
        else:
            # 7/14/28일: <10 수정 + 과학적 표기법
            value = self._fix_less_than_10(value)
            value = self._normalize_scientific(value)
            
            # 7일차 추가 보정: 애매한 경우 <10^2 고려
            if day_column == '7일':
                value = self._fix_7day_ambiguous(value, original_value)
        
        # 3단계: 보정 로그 (변경된 경우만)
        if value != original_value and value != '':
            logger.info(f"  🔧 보정 [{day_column}]: '{original_value}' → '{value}'")
        
        return value
    
    def _fix_7day_ambiguous(self, value: str, original: str) -> str:
        """
        7일차 애매한 값 추가 보정
        
        ISO 11930 실무 기준:
        - 7일차는 일반적으로 <10^2 (100 미만) 기준 사용
        - <10 (10 미만)은 매우 엄격한 기준으로 드물게 사용
        
        전략:
        - 명확한 <10 패턴만 <10 유지
        - 애매한 오인식 패턴은 <10^2로 보정
        
        Args:
            value: 보정된 값
            original: 원본 값
            
        Returns:
            최종 값
        """
        # 이미 지수 형태면 그대로
        if '^' in value:
            return value
        
        # <10이 아니면 그대로
        if value != '<10':
            return value
        
        # 원본이 명확한 <10 패턴이면 그대로 유지
        clear_less_10_patterns = [
            '< 10',   # 공백 포함
            '<10',    # 공백 없음
            '< 10"',  # 따옴표 포함
            '<10"',
            '< 10\''  # 작은따옴표
        ]
        
        original_clean = original.strip()
        
        for pattern in clear_less_10_patterns:
            if original_clean == pattern or original_clean == pattern.replace(' ', ''):
                # 명확한 <10 → 유지
                return '<10'
        
        # 애매한 오인식 패턴
        # (40, CIO 등 - 원본 의도가 <10인지 <10^2인지 불명확)
        # → 실무에서는 7일차 <10^2가 더 일반적이므로 <10^2로 보정
        ambiguous_patterns = ['40', '40°', '40€', 'CIO', 'CIÒ', 'C10', '410', '90']
        
        for pattern in ambiguous_patterns:
            if pattern in original_clean:
                # 애매한 패턴 → <10^2 (실무 기준)
                logger.info(f"  ℹ️ 7일차 실무 보정: '{original}' → '<10^2' (일반 기준)")
                return '<10^2'
        
        # 기타 알 수 없는 패턴 → <10 유지 (보수적)
        return '<10'
    
    def _extract_judgment(self, value: str) -> str:
        """판정 값 추출"""
        if not value:
            return "적합"
        
        value = value.strip().upper()
        
        # 부적합 패턴
        if any(char in value for char in ['X', '×', 'V', '부적합']):
            return '부적합'
        
        return '적합'
    
    @staticmethod
    def convert_to_log(cfu_value: str) -> str:
        """CFU → Log 변환"""
        if not cfu_value:
            return ""
        
        try:
            if '<' in cfu_value:
                if '10^' in cfu_value:
                    exp_match = re.search(r'<10\^(\d+)', cfu_value)
                    if exp_match:
                        return f"<{exp_match.group(1)}.0"
                return "<1.0"
            
            exp_match = re.match(r'([0-9.]+)×10\^(\d+)', cfu_value)
            if exp_match:
                base = float(exp_match.group(1))
                exp = int(exp_match.group(2))
                log_value = exp + math.log10(base)
                return round(log_value, 1)
            
            try:
                num = float(cfu_value)
                return round(math.log10(num), 1)
            except ValueError:
                pass
            
            return cfu_value
            
        except Exception as e:
            logger.warning(f"Log 변환 실패: {cfu_value}, 오류: {e}")
            return cfu_value


def process_preservation_page(pdf_bytes: bytes, page_index: int, excel_path: str = None) -> dict:
    """
    보존력 시험 페이지 처리 (Azure OCR 기반)
    
    Args:
        pdf_bytes: PDF 바이트 데이터
        page_index: 페이지 인덱스
        excel_path: TestResult_PROGRESS.xlsx 파일 경로 (선택사항)
    
    환경변수 DEBUG_MODE=1 설정 시 상세 디버깅 정보 출력
    """
    result = {
        'success': False,
        'data': [],
        'date_info': {},
        'message': ''
    }
    
    temp_image_path = None
    
    # 🐛 디버그 모드 확인
    debug_mode = os.getenv('DEBUG_MODE', '0') == '1'
    
    try:
        # 0. Excel 진행서 데이터 로드 (선택사항)
        progress_data = {}
        if excel_path:
            progress_data = load_progress_excel(excel_path)
        
        # 1. DRM 처리
        drm_success, processed_bytes, drm_message = PDFProcessor.process_drm_if_needed(pdf_bytes)
        if not drm_success:
            result['message'] = drm_message
            return result
        
        logger.info(f"📄 DRM 처리: {drm_message}")
        
        # 2. 이미지 렌더링
        img_bytes = PDFProcessor.render_page_image(processed_bytes, page_index, zoom=2.0)
        if not img_bytes:
            result['message'] = "이미지 렌더링 실패"
            return result
        
        # 3. 임시 파일 저장
        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as temp_file:
            temp_image_path = temp_file.name
            temp_file.write(img_bytes)
        
        logger.info(f"💾 임시 이미지 저장: {temp_image_path}")
        
        # 4. Azure OCR (디버그 모드 전달)
        ocr = PreservationTestOCR(debug_mode=debug_mode)
        test_data = ocr.extract_preservation_test_table(temp_image_path)
        
        if not test_data or not test_data.get('data'):
            result['message'] = "데이터 추출 실패"
            return result
        
        # 5. Excel 데이터와 매칭 (진행서 데이터가 있는 경우)
        if progress_data:
            logger.info(f"\n🔗 진행서 데이터 매칭 시작")
            matched_count = 0
            for item in test_data['data']:
                prescription = item.get('prescription_number', '')
                if prescription and prescription in progress_data:
                    # 추가 정보 병합
                    item['product_name'] = progress_data[prescription]['product_name']
                    item['formulation'] = progress_data[prescription]['formulation']
                    item['preservative_info'] = progress_data[prescription]['preservative_info']
                    matched_count += 1
                    logger.info(f"  ✅ 매칭 성공: {prescription} → {item['product_name']}")
                else:
                    # 매칭 실패 시 빈 값
                    item['product_name'] = ''
                    item['formulation'] = ''
                    item['preservative_info'] = ''
                    if prescription:
                        logger.warning(f"  ⚠️ 매칭 실패: {prescription}")
            
            logger.info(f"  📊 매칭 결과: {matched_count}/{len(test_data['data'])}")
        
        # 6. 결과 포맷팅
        result['success'] = True
        result['data'] = test_data['data']
        result['date_info'] = test_data['date_info']
        result['message'] = f"{len(test_data['data'])}개 균주 데이터 추출 완료"
        
        logger.info(f"✅ OCR 성공: {len(test_data['data'])}개 균주")
        
        return result
        
    except Exception as e:
        logger.error(f"❌ 처리 오류: {e}")
        import traceback
        traceback.print_exc()
        result['message'] = str(e)
        return result
    
    finally:
        # 임시 파일 정리
        if temp_image_path and os.path.exists(temp_image_path):
            try:
                os.remove(temp_image_path)
                logger.info(f"🗑️ 임시 파일 삭제: {temp_image_path}")
            except Exception as e:
                logger.warning(f"⚠️ 임시 파일 삭제 실패: {e}")


class PreservationExcelSaver:
    """보존력 시험 Excel 저장 (템플릿 기반)"""
    
    DEFAULT_TEMPLATE = "TestResult_OCR_v1.xlsx"
    DEFAULT_PROGRESS_FILE = "TestResult_PROGRESS.xlsx"
    
    def __init__(self, output_path: str, template_file: str = None, progress_file: str = None):
        self.output_path = output_path
        self.template_file = template_file or self.DEFAULT_TEMPLATE
        self.progress_file = progress_file or self.DEFAULT_PROGRESS_FILE
        
        # 템플릿 파일 확인
        if not os.path.exists(self.template_file):
            logger.warning(f"⚠️ 템플릿 파일을 찾을 수 없습니다: {self.template_file}")
            self.template_file = None
        else:
            logger.info(f"✅ 템플릿 파일 확인: {self.template_file}")
        
        # TestResult_PROGRESS.xlsx 파일 확인 및 제품 정보 로드
        self.product_info_dict = {}
        if os.path.exists(self.progress_file):
            logger.info(f"✅ TestResult_PROGRESS.xlsx 파일 확인: {self.progress_file}")
            self.product_info_dict = self._load_product_info_from_progress_file()
        else:
            logger.warning(f"⚠️ TestResult_PROGRESS.xlsx 파일을 찾을 수 없습니다: {self.progress_file}")
        
        # 파일이 없으면 초기화
        if not os.path.exists(self.output_path):
            self._initialize_excel()
    
    def _initialize_excel(self):
        """Excel 파일 초기화"""
        try:
            if self.template_file and os.path.exists(self.template_file):
                import shutil
                from openpyxl import load_workbook
                
                shutil.copy2(self.template_file, self.output_path)
                
                workbook = load_workbook(self.output_path)
                if len(workbook.sheetnames) > 0:
                    first_sheet = workbook[workbook.sheetnames[0]]
                    first_sheet.title = "TEMPLATE_BASE"
                
                workbook.save(self.output_path)
                workbook.close()
                
                logger.info(f"✅ 템플릿 기반 Excel 초기화 완료")
            else:
                from openpyxl import Workbook
                wb = Workbook()
                wb.remove(wb.active)
                wb.save(self.output_path)
                wb.close()
                
                logger.warning(f"⚠️ 템플릿 없이 빈 Excel 파일 생성")
            
            return True
        except Exception as e:
            logger.error(f"❌ Excel 초기화 실패: {e}")
            return False
    
    def _load_product_info_from_progress_file(self):
        """
        TestResult_PROGRESS.xlsx에서 제품 정보 로드
        
        컬럼 구조:
        - 제품명 (A열, 1번째)
        - 처방번호 (B열, 2번째)
        - 제형 (C열, 3번째)
        - 고시미등록방부보조성분함량 (E열, 5번째)
        
        Returns:
            dict: {처방번호: {'제품명': ..., '제형': ..., '고시미등록방부보조성분함량': ...}}
        """
        product_dict = {}
        
        try:
            from openpyxl import load_workbook
            
            if not os.path.exists(self.progress_file):
                logger.warning(f"⚠️ TestResult_PROGRESS.xlsx 파일 없음: {self.progress_file}")
                return product_dict
            
            logger.info(f"📖 TestResult_PROGRESS.xlsx 파일 읽기 시작: {self.progress_file}")
            
            workbook = load_workbook(self.progress_file, read_only=True, data_only=True)
            
            # 첫 번째 시트 사용
            if len(workbook.sheetnames) == 0:
                logger.warning("⚠️ 시트가 없습니다")
                workbook.close()
                return product_dict
            
            sheet = workbook[workbook.sheetnames[0]]
            logger.info(f"📄 시트 이름: {sheet.title}")
            
            # 데이터 읽기 (헤더 스킵, 2번째 행부터)
            row_count = 0
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row or len(row) < 5:
                    continue
                
                # 컬럼 인덱스: 0부터 시작
                product_name = str(row[0]).strip() if row[0] else ''
                prescription_number = str(row[1]).strip() if row[1] else ''
                formulation_type = str(row[2]).strip() if row[2] else ''
                # row[3]은 소속 (사용 안 함)
                unregistered_preservatives = str(row[4]).strip() if row[4] else ''
                
                # 처방번호가 있어야 매칭 가능
                if not prescription_number:
                    continue
                
                # 딕셔너리에 저장
                product_dict[prescription_number] = {
                    '제품명': product_name,
                    '제형': formulation_type,
                    '고시미등록방부보조성분함량': unregistered_preservatives
                }
                
                row_count += 1
            
            workbook.close()
            
            logger.info(f"✅ 제품 정보 로드 완료: {row_count}개 제품")
            logger.info(f"📋 처방번호 목록 (일부): {list(product_dict.keys())[:5]}...")
            
            return product_dict
            
        except Exception as e:
            logger.error(f"❌ 제품 정보 로드 실패: {e}")
            import traceback
            traceback.print_exc()
            return product_dict
    
    def add_test_data(self, test_data, date_info=None):
        """
        테스트 데이터를 Excel에 추가
        
        Args:
            test_data: DataFrame 또는 딕셔너리 리스트
            date_info: 날짜 정보 딕셔너리
        """
        try:
            from openpyxl import load_workbook
            import pandas as pd
            
            # DataFrame으로 변환
            if isinstance(test_data, pd.DataFrame):
                df = test_data
            elif isinstance(test_data, list):
                df = pd.DataFrame(test_data)
            else:
                logger.error("❌ 지원하지 않는 데이터 형식")
                return False
            
            if df.empty:
                logger.warning("⚠️ 빈 데이터")
                return False
            
            # 시험번호 확인
            if 'test_number' not in df.columns:
                logger.error("❌ test_number 컬럼이 없습니다")
                return False
            
            # 시험번호별로 그룹핑
            test_numbers = df['test_number'].dropna().unique()
            
            if len(test_numbers) == 0:
                logger.warning("⚠️ 유효한 시험번호가 없습니다")
                return False
            
            logger.info(f"📋 {len(test_numbers)}개 시험번호 발견: {list(test_numbers)}")
            
            # Excel 파일 로드
            workbook = load_workbook(self.output_path)
            
            success_count = 0
            
            # 각 시험번호별로 처리
            for test_number in test_numbers:
                if not test_number or str(test_number).strip() == '':
                    continue
                
                # 해당 시험번호의 데이터만 추출
                df_subset = df[df['test_number'] == test_number].copy()
                
                if df_subset.empty:
                    continue
                
                logger.info(f"🔄 {test_number} 처리 중... ({len(df_subset)}개 행)")
                
                # 시트명 설정
                sheet_name = str(test_number)
                
                # 🔧 기존 시트가 있으면 삭제 (업데이트를 위해)
                if sheet_name in workbook.sheetnames:
                    del workbook[sheet_name]
                    logger.info(f"🔄 기존 시트 삭제 (업데이트): {sheet_name}")
                
                # 템플릿 시트 복사하여 새 시트 생성
                if "TEMPLATE_BASE" in workbook.sheetnames:
                    template_sheet = workbook["TEMPLATE_BASE"]
                    new_sheet = workbook.copy_worksheet(template_sheet)
                    new_sheet.title = sheet_name
                    logger.info(f"✅ 시트 생성 완료: {sheet_name}")
                else:
                    new_sheet = workbook.create_sheet(title=sheet_name)
                    logger.warning(f"⚠️ 템플릿 없이 빈 시트 생성: {sheet_name}")
                
                # 데이터 매핑
                self._map_data_to_sheet(new_sheet, df_subset, date_info)
                
                success_count += 1
            
            # 즉시 저장
            workbook.save(self.output_path)
            workbook.close()
            
            logger.info(f"💾 Excel 저장 완료: {success_count}개 시트")
            return True
            
        except Exception as e:
            logger.error(f"❌ Excel 저장 실패: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _map_data_to_sheet(self, worksheet, df, date_info):
        """
        데이터를 템플릿 시트에 매핑
        
        템플릿 구조 (사용자 정의):
        
        원본 영역 (상단):
        - AA3: 시험번호 (원본)
        - E3: 제품명 (원본)
        - E4: 처방번호 (원본)
        - Y5: 제형 (원본)
        - E6: 고시미등록방부보조성분함량 (원본)
        - D30: 최종판정 (원본)
        - I19, L19, O19, R19: 날짜 (0, 7, 14, 28일) - 원본
        - J20-U24: 균주별 CFU 값 (원본)
        
        로그 영역 (하단):
        - AA33: 시험번호 (로그)
        - E33: 제품명 (로그)
        - E34: 처방번호 (로그)
        - Y35: 제형 (로그)
        - E36: 고시미등록방부보조성분함량 (로그)
        - D60: 최종판정 (로그)
        - I49, L49, O49, R49: 날짜 - 로그
        - J50-S54: 균주별 Log 값
        """
        try:
            if df.empty:
                return
            
            # 🔧 사용자 정의 매핑
            # 시험번호 매핑 (AA3: 원본, AA33: 로그)
            test_number = df.iloc[0].get('test_number', '')
            if test_number:
                worksheet['AA3'] = test_number    # 원본
                worksheet['AA33'] = test_number   # 로그
                logger.info(f"📝 시험번호 매핑: AA3, AA33 = {test_number}")
            
            # 처방번호 매핑 (E4: 원본, E34: 로그)
            prescription_number = df.iloc[0].get('prescription_number', '')
            if prescription_number:
                worksheet['E4'] = prescription_number    # 원본
                worksheet['E34'] = prescription_number   # 로그
                logger.info(f"📝 처방번호 매핑: E4, E34 = {prescription_number}")
            
            # 🆕 OCR 결과에서 추가 정보 직접 읽기 (우선순위)
            product_name_from_ocr = df.iloc[0].get('product_name', '')
            formulation_from_ocr = df.iloc[0].get('formulation', '')
            preservative_from_ocr = df.iloc[0].get('preservative_info', '')
            
            # 제품명 매핑 (E3: 원본, E33: 로그)
            product_name = product_name_from_ocr or (
                self.product_info_dict.get(prescription_number, {}).get('제품명', '') 
                if prescription_number in self.product_info_dict else ''
            )
            if product_name:
                worksheet['E3'] = product_name    # 원본
                worksheet['E33'] = product_name   # 로그
                logger.info(f"📝 제품명 매핑: E3, E33 = {product_name}")
            
            # 제형 매핑 (Y5: 원본, Y35: 로그)
            formulation = formulation_from_ocr or (
                self.product_info_dict.get(prescription_number, {}).get('제형', '')
                if prescription_number in self.product_info_dict else ''
            )
            if formulation:
                worksheet['Y5'] = formulation     # 원본
                worksheet['Y35'] = formulation    # 로그
                logger.info(f"📝 제형 매핑: Y5, Y35 = {formulation}")
            
            # 고시미등록방부보조성분함량 매핑 (E6: 원본, E36: 로그)
            preservative_info = preservative_from_ocr or (
                self.product_info_dict.get(prescription_number, {}).get('고시미등록방부보조성분함량', '')
                if prescription_number in self.product_info_dict else ''
            )
            if preservative_info:
                worksheet['E6'] = preservative_info    # 원본
                worksheet['E36'] = preservative_info   # 로그
                logger.info(f"📝 고시미등록방부보조성분함량 매핑: E6, E36 = {preservative_info}")
            
            # 최종판정 매핑 (D30: 원본, D60: 로그)
            final_judgment = df.iloc[0].get('final_judgment', '')
            if final_judgment:
                worksheet['D30'] = final_judgment    # 원본
                worksheet['D60'] = final_judgment    # 로그
                logger.info(f"📝 최종판정 매핑: D30, D60 = {final_judgment}")
            else:
                # 빈 값이면 공란으로
                worksheet['D30'] = ''
                worksheet['D60'] = ''
                logger.info(f"📝 최종판정 매핑: D30, D60 = (공란)")
            
            # 날짜 정보 매핑
            if date_info:
                date_list = [
                    date_info.get('date_0', ''),
                    date_info.get('date_7', ''),
                    date_info.get('date_14', ''),
                    date_info.get('date_28', '')
                ]
                
                if len(date_list) >= 4:
                    date_positions_original = ['I19', 'L19', 'O19', 'R19']
                    date_positions_log = ['I49', 'L49', 'O49', 'R49']
                    
                    for i, date_val in enumerate(date_list[:4]):
                        if date_val:
                            worksheet[date_positions_original[i]] = date_val
                            worksheet[date_positions_log[i]] = date_val
                    
                    logger.info(f"📅 날짜 정보 매핑: {date_list}")
            
            # 균주별 CFU 데이터 매핑
            strain_mapping = {
                'E.coli': 'E.coli',
                'P.aeruginosa': 'P.aeruginosa',
                'S.aureus': 'S.aureus',
                'C.albicans': 'C.albicans',
                'A.brasiliensis': 'A.brasiliensis'
            }
            
            original_positions = {
                'E.coli': ['J20', 'M20', 'P20', 'S20', 'U20'],
                'P.aeruginosa': ['J21', 'M21', 'P21', 'S21', 'U21'],
                'S.aureus': ['J22', 'M22', 'P22', 'S22', 'U22'],
                'C.albicans': ['J23', 'M23', 'P23', 'S23', 'U23'],
                'A.brasiliensis': ['J24', 'M24', 'P24', 'S24', 'U24']
            }
            
            log_positions = {
                'E.coli': ['J50', 'M50', 'P50', 'S50'],
                'P.aeruginosa': ['J51', 'M51', 'P51', 'S51'],
                'S.aureus': ['J52', 'M52', 'P52', 'S52'],
                'C.albicans': ['J53', 'M53', 'P53', 'S53'],
                'A.brasiliensis': ['J54', 'M54', 'P54', 'S54']
            }
            
            mapped_count = 0
            for _, row in df.iterrows():
                strain = row.get('strain', '')
                if not strain:
                    continue
                
                mapped_strain = strain_mapping.get(strain, strain)
                
                if mapped_strain in original_positions:
                    # 원본 CFU 값
                    positions = original_positions[mapped_strain]
                    worksheet[positions[0]] = row.get('cfu_0day', '')
                    worksheet[positions[1]] = row.get('cfu_7day', '')
                    worksheet[positions[2]] = row.get('cfu_14day', '')
                    worksheet[positions[3]] = row.get('cfu_28day', '')
                    worksheet[positions[4]] = row.get('judgment', '')
                    
                    # Log 값
                    log_pos = log_positions[mapped_strain]
                    worksheet[log_pos[0]] = PreservationTestOCR.convert_to_log(row.get('cfu_0day', ''))
                    worksheet[log_pos[1]] = PreservationTestOCR.convert_to_log(row.get('cfu_7day', ''))
                    worksheet[log_pos[2]] = PreservationTestOCR.convert_to_log(row.get('cfu_14day', ''))
                    worksheet[log_pos[3]] = PreservationTestOCR.convert_to_log(row.get('cfu_28day', ''))
                    
                    mapped_count += 1
                    logger.info(f"🦠 {mapped_strain} 데이터 매핑 완료")
            
            logger.info(f"✅ 총 {mapped_count}개 균주 데이터 매핑 완료")
            
        except Exception as e:
            logger.error(f"❌ 데이터 매핑 실패: {e}")
            import traceback
            traceback.print_exc()
    
    def get_sheet_list(self):
        """시트 목록 반환"""
        try:
            from openpyxl import load_workbook
            
            if os.path.exists(self.output_path):
                workbook = load_workbook(self.output_path, read_only=True)
                sheet_names = workbook.sheetnames
                workbook.close()
                
                filtered_names = [name for name in sheet_names if name != "TEMPLATE_BASE"]
                return filtered_names
            else:
                return []
        except Exception as e:
            logger.error(f"❌ 시트 목록 조회 실패: {e}")
            return []
    
    def get_excel_bytes(self):
        """Excel 바이트 반환"""
        try:
            if os.path.exists(self.output_path):
                with open(self.output_path, 'rb') as f:
                    return f.read()
        except:
            pass
        return None
    
    def get_statistics(self):
        """통계 반환"""
        try:
            from openpyxl import load_workbook
            if os.path.exists(self.output_path):
                wb = load_workbook(self.output_path, read_only=True)
                total_sheets = len(wb.sheetnames)
                test_sheets = len([name for name in wb.sheetnames if name != "TEMPLATE_BASE"])
                wb.close()
                
                file_size = os.path.getsize(self.output_path)
                return {
                    'total_sheets': total_sheets,
                    'test_sheets': test_sheets,
                    'file_size': file_size,
                    'file_size_mb': round(file_size / (1024 * 1024), 2)
                }
        except:
            pass
        
        return {'total_sheets': 0, 'test_sheets': 0, 'file_size': 0, 'file_size_mb': 0}